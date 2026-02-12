"""
Reusable Excel cell/row formatting helpers.
"""
from __future__ import annotations

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.excel.styles import (
    HEADER_FONT, HEADER_FILL, HEADER_BORDER,
    DATA_FONT, TOTAL_FONT,
    THIN_BORDER, TOTAL_BORDER,
    ALTERNATE_FILL, TOTAL_FILL,
    KPI_VALUE_FONT, KPI_LABEL_FONT,
    CENTER, LEFT, RIGHT,
    HIGHLIGHT_FILLS,
)


# ---------------------------------------------------------------------------
# Header row
# ---------------------------------------------------------------------------

def format_header_row(ws: Worksheet, row_num: int, num_cols: int) -> None:
    """Apply header styling to an entire row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = HEADER_BORDER


# ---------------------------------------------------------------------------
# Data cell
# ---------------------------------------------------------------------------

def format_data_cell(
    ws: Worksheet,
    row_num: int,
    col_num: int,
    value,
    col_type: str = "text",
    is_total: bool = False,
    highlight: str | None = None,
) -> None:
    """Write and format a single data cell."""
    cell = ws.cell(row=row_num, column=col_num)
    cell.value = value
    cell.font = TOTAL_FONT if is_total else DATA_FONT
    cell.border = TOTAL_BORDER if is_total else THIN_BORDER
    cell.alignment = RIGHT if col_type in ("currency", "number", "percent", "decimal") else LEFT

    if col_type == "currency":
        cell.number_format = '"$"#,##0.00'
    elif col_type == "percent":
        cell.number_format = '0.0"%"'
    elif col_type == "number":
        cell.number_format = "#,##0"
    elif col_type == "decimal":
        cell.number_format = "0.0"

    if highlight and highlight in HIGHLIGHT_FILLS:
        cell.fill = HIGHLIGHT_FILLS[highlight]
    elif is_total:
        cell.fill = TOTAL_FILL
    elif row_num % 2 == 0:
        cell.fill = ALTERNATE_FILL


# ---------------------------------------------------------------------------
# Auto column width
# ---------------------------------------------------------------------------

def auto_column_width(ws: Worksheet, min_width: int = 10, max_width: int = 55) -> None:
    """Auto-fit column widths based on content length."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass
        adjusted = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted


# ---------------------------------------------------------------------------
# KPI card
# ---------------------------------------------------------------------------

def add_kpi_card(
    ws: Worksheet,
    row: int,
    col: int,
    value,
    label: str,
    format_type: str = "currency",
) -> None:
    """Write a large KPI value + small label below it."""
    value_cell = ws.cell(row=row, column=col)
    label_cell = ws.cell(row=row + 1, column=col)

    value_cell.value = value
    value_cell.font = KPI_VALUE_FONT
    value_cell.alignment = CENTER

    if format_type == "currency":
        value_cell.number_format = '"$"#,##0'
    elif format_type == "percent":
        value_cell.number_format = '0.0"%"'
    elif format_type == "number":
        value_cell.number_format = "#,##0"
    elif format_type == "decimal":
        value_cell.number_format = "0.0"
    # format_type == "text" → no number_format

    label_cell.value = label
    label_cell.font = KPI_LABEL_FONT
    label_cell.alignment = CENTER
