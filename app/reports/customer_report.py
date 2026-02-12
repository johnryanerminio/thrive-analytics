"""
Master Customer Insights Report — Period-specific metrics with segmentation.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd

from app.data.store import DataStore
from app.data.schemas import PeriodFilter
from app.analytics.common import sanitize_for_json
from app.analytics.customers import customer_metrics, customer_summary, segment_summary, top_customers
from app.excel.writer import ExcelWriter
from app.excel.styles import SECTION_FONT
from app.excel.formatters import add_kpi_card


TOP_CUST_COLS = [
    ("rank", "number", "Rank"),
    ("customer_name", "text", "Customer Name"),
    ("primary_store", "text", "Primary Store"),
    ("segment", "text", "Segment"),
    ("is_loyal", "text", "Loyal"),
    ("total_spent", "currency", "Period Spend"),
    ("transactions", "number", "Transactions"),
    ("avg_transaction", "currency", "Avg Transaction"),
    ("total_discounts", "currency", "Discounts"),
    ("discount_rate", "percent", "Discount Rate"),
]


def generate_json(store: DataStore, period: PeriodFilter | None = None) -> dict:
    sales_df = store.get_sales(period)
    date_range = store.date_range(period)
    summary = customer_summary(sales_df, store.cust_attr_df)
    cust_df = customer_metrics(sales_df, store.cust_attr_df)

    segments = segment_summary(cust_df, summary["total_revenue"])
    top = top_customers(cust_df, 50)

    # Top 50 by store
    by_store = {}
    for s in sorted(cust_df["primary_store"].dropna().unique()):
        if s == "Unknown":
            continue
        store_cust = cust_df[cust_df["primary_store"] == s].nlargest(50, "total_spent").copy()
        if not store_cust.empty:
            store_cust["rank"] = range(1, len(store_cust) + 1)
            by_store[s] = store_cust.fillna(0).to_dict("records")

    return sanitize_for_json({
        "date_range": date_range,
        "summary": summary,
        "segments": segments,
        "top_customers": top,
        "by_store": by_store,
    })


def generate_excel(
    store: DataStore,
    output_path: str | Path,
    period: PeriodFilter | None = None,
) -> Path:
    data = generate_json(store, period)
    ew = ExcelWriter()
    s = data["summary"]
    dr = data["date_range"]

    # Executive Summary
    ws = ew.add_sheet("Executive Summary")
    ew.write_title(ws, "THRIVE CANNABIS",
                   f"Customer Insights Report  |  {dr}  |  Generated {pd.Timestamp.now():%B %d, %Y}")

    from app.excel.styles import RED
    from openpyxl.styles import Font
    ws.cell(row=3, column=1).value = f"DATA PERIOD: {dr}"
    ws.cell(row=3, column=1).font = Font(name="Calibri", size=11, bold=True, color=RED)

    row = ew.write_section(ws, 5, "CUSTOMER OVERVIEW")
    row = ew.write_kpi_row(ws, row, [
        (s["total_customers"], "UNIQUE CUSTOMERS", "number"),
        (s["total_revenue"], "TOTAL REVENUE", "currency"),
        (s["revenue_per_customer"], "REVENUE/CUSTOMER", "currency"),
        (s["loyalty_rate"], "LOYALTY RATE", "percent"),
    ])

    row = ew.write_section(ws, row, "TRANSACTION METRICS")
    ew.write_kpi_row(ws, row, [
        (s["total_transactions"], "TOTAL TRANSACTIONS", "number"),
        (s["avg_transaction"], "AVG TRANSACTION", "currency"),
        (s["total_discounts"], "TOTAL DISCOUNTS", "currency"),
        (s["discount_rate"], "DISCOUNT RATE", "percent"),
    ])

    # Segments
    ws2 = ew.add_sheet("Customer Segments")
    ew.write_table(ws2, 1, [
        ("segment", "text", "Segment"),
        ("customers", "number", "Customers"),
        ("pct_of_cust", "percent", "% of Customers"),
        ("total_revenue", "currency", "Total Revenue"),
        ("pct_of_rev", "percent", "% of Revenue"),
        ("rev_per_cust", "currency", "Rev/Customer"),
        ("total_discounts", "currency", "Total Discounts"),
        ("discount_rate", "percent", "Discount Rate"),
    ], data["segments"])

    # High Value Customers
    ws3 = ew.add_sheet("High Value Customers")
    ws3.cell(row=1, column=1).value = f"Top 50 Customers by Spend During {dr}"
    ws3.cell(row=1, column=1).font = SECTION_FONT
    ew.write_table(ws3, 3, TOP_CUST_COLS, data["top_customers"],
                   highlight_fn=lambda i, r: "gold" if i < 10 else None)

    # By store tabs
    for store_name, cust_list in data["by_store"].items():
        if not cust_list:
            continue
        short = store_name.replace("Thrive ", "").replace("Cannabis ", "")[:12]
        ws_s = ew.add_sheet(f"Top 50 - {short}")
        ws_s.cell(row=1, column=1).value = f"{store_name} - Top 50 ({dr})"
        ws_s.cell(row=1, column=1).font = SECTION_FONT
        ew.write_table(ws_s, 3, TOP_CUST_COLS, cust_list,
                       highlight_fn=lambda i, r: "gold" if i < 10 else None)

    return ew.save(output_path)
