"""
ExcelWriter — high-level helpers for building styled Excel workbooks.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.excel.styles import (
    TITLE_FONT, SUBTITLE_FONT, SECTION_FONT,
    KPI_VALUE_FONT, KPI_LABEL_FONT,
    POSITIVE_KPI_FONT, NEGATIVE_KPI_FONT,
    INSIGHT_TITLE_FONT, INSIGHT_BODY_FONT,
    REC_TITLE_FONT, REC_BODY_FONT,
    LEGEND_BOLD_FONT, DATA_FONT,
    LIGHT_GREEN_FILL, THIN_BORDER, WRAP,
    CENTER,
)
from app.excel.formatters import (
    format_header_row,
    format_data_cell,
    auto_column_width,
    add_kpi_card,
)


ColSpec = tuple[str, str, str]  # (key, col_type, label)


class ExcelWriter:
    """Fluent builder for styled Excel workbooks."""

    def __init__(self) -> None:
        self.wb = Workbook()
        self._first_sheet = True

    # ------------------------------------------------------------------
    # Sheet management
    # ------------------------------------------------------------------

    def add_sheet(self, title: str) -> Worksheet:
        """Create a new worksheet (re-uses the default sheet for the first call)."""
        if self._first_sheet:
            ws = self.wb.active
            ws.title = title
            self._first_sheet = False
        else:
            ws = self.wb.create_sheet(title=title)
        return ws

    # ------------------------------------------------------------------
    # Title block
    # ------------------------------------------------------------------

    def write_title(
        self,
        ws: Worksheet,
        title: str,
        subtitle: str,
        merge_cols: int = 8,
    ) -> int:
        """Write title + subtitle rows. Returns next available row."""
        ws.cell(row=1, column=1).value = title
        ws.cell(row=1, column=1).font = TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_cols)

        ws.cell(row=2, column=1).value = subtitle
        ws.cell(row=2, column=1).font = SUBTITLE_FONT
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=merge_cols)

        # Default column widths
        for col in range(1, merge_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        return 4  # next row

    def write_section(self, ws: Worksheet, row: int, title: str) -> int:
        """Write a section header. Returns next row."""
        ws.cell(row=row, column=1).value = title
        ws.cell(row=row, column=1).font = SECTION_FONT
        return row + 2

    # ------------------------------------------------------------------
    # KPI cards
    # ------------------------------------------------------------------

    def write_kpi_row(
        self,
        ws: Worksheet,
        row: int,
        kpis: list[tuple],  # [(value, label, format_type), ...]
        start_col: int = 1,
        col_spacing: int = 2,
    ) -> int:
        """Write a row of KPI cards. Returns next row (row + 2)."""
        col = start_col
        for value, label, fmt in kpis:
            add_kpi_card(ws, row, col, value, label, fmt)
            col += col_spacing
        return row + 3

    def write_delta_kpi(
        self,
        ws: Worksheet,
        row: int,
        col: int,
        value: float,
        label: str,
    ) -> None:
        """Write a KPI that's green when positive, red when negative."""
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.font = POSITIVE_KPI_FONT if value >= 0 else NEGATIVE_KPI_FONT
        cell.number_format = '+0.0;-0.0;0.0'
        cell.alignment = CENTER

        lbl = ws.cell(row=row + 1, column=col)
        lbl.value = label
        lbl.font = KPI_LABEL_FONT
        lbl.alignment = CENTER

    # ------------------------------------------------------------------
    # Data tables
    # ------------------------------------------------------------------

    def write_table(
        self,
        ws: Worksheet,
        start_row: int,
        columns: list[ColSpec],
        data: list[dict] | pd.DataFrame,
        highlight_fn=None,
        freeze: bool = True,
        show_total: bool = False,
        total_label: str = "TOTAL",
    ) -> int:
        """Write a full table with headers + data rows.

        highlight_fn(row_idx, row_data) -> str|None  e.g. 'gold', 'warning'

        Returns the row number after the last data row.
        """
        # Headers
        for col_num, (_, _, label) in enumerate(columns, 1):
            ws.cell(row=start_row, column=col_num).value = label
        format_header_row(ws, start_row, len(columns))

        # Data rows
        if isinstance(data, pd.DataFrame):
            rows = data.to_dict("records")
        else:
            rows = data

        row = start_row + 1
        for idx, row_data in enumerate(rows):
            for col_num, (key, col_type, _) in enumerate(columns, 1):
                val = row_data.get(key, 0)
                if pd.isna(val):
                    val = 0
                hl = highlight_fn(idx, row_data) if highlight_fn else None
                format_data_cell(ws, row, col_num, val, col_type, highlight=hl)
            row += 1

        # Total row
        if show_total and rows:
            df_rows = pd.DataFrame(rows)
            format_data_cell(ws, row, 1, total_label, "text", is_total=True)
            for col_num, (key, col_type, _) in enumerate(columns[1:], 2):
                if col_type in ("currency", "number"):
                    val = df_rows[key].sum() if key in df_rows.columns else 0
                    format_data_cell(ws, row, col_num, val, col_type, is_total=True)
                else:
                    format_data_cell(ws, row, col_num, "", "text", is_total=True)
            row += 1

        auto_column_width(ws)
        if freeze:
            ws.freeze_panes = f"A{start_row + 1}"

        return row

    # ------------------------------------------------------------------
    # Insight / recommendation blocks
    # ------------------------------------------------------------------

    def write_insight(self, ws: Worksheet, row: int, title: str, body: str, merge_cols: int = 8) -> int:
        """Write a key insight block. Returns next row."""
        ws.cell(row=row, column=1).value = title
        ws.cell(row=row, column=1).font = INSIGHT_TITLE_FONT
        ws.cell(row=row + 1, column=1).value = body
        ws.cell(row=row + 1, column=1).font = INSIGHT_BODY_FONT
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=merge_cols)
        return row + 3

    def write_recommendations(
        self,
        ws: Worksheet,
        start_row: int,
        recs: list[dict],
        merge_cols: int = 6,
    ) -> int:
        """Write recommendation blocks. Returns next row."""
        row = start_row
        severity_icons = {"red": "🔴", "yellow": "🟡", "green": "🟢", "info": "💰"}
        for rec in recs:
            icon = severity_icons.get(rec.get("severity", "info"), "💡")
            ws.cell(row=row, column=1).value = f"{icon} {rec['title']}"
            ws.cell(row=row, column=1).font = REC_TITLE_FONT
            ws.cell(row=row + 1, column=1).value = rec.get("detail", "")
            ws.cell(row=row + 1, column=1).font = REC_BODY_FONT
            ws.cell(row=row + 1, column=1).alignment = WRAP
            ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=merge_cols)
            if rec.get("action"):
                ws.cell(row=row + 2, column=1).value = f"  → {rec['action']}"
                ws.cell(row=row + 2, column=1).font = DATA_FONT
                ws.merge_cells(start_row=row + 2, start_column=1, end_row=row + 2, end_column=merge_cols)
                row += 4
            else:
                row += 3
        return row

    def write_legend(self, ws: Worksheet, start_row: int, items: list[tuple[str, str]]) -> int:
        """Write a classification legend table. Returns next row."""
        ws.cell(row=start_row, column=1).value = "Category"
        ws.cell(row=start_row, column=2).value = "What It Includes"
        format_header_row(ws, start_row, 2)

        row = start_row + 1
        for cat, desc in items:
            c1 = ws.cell(row=row, column=1)
            c1.value = cat
            c1.font = LEGEND_BOLD_FONT
            c1.fill = LIGHT_GREEN_FILL
            c1.border = THIN_BORDER

            c2 = ws.cell(row=row, column=2)
            c2.value = desc
            c2.font = DATA_FONT
            c2.fill = LIGHT_GREEN_FILL
            c2.border = THIN_BORDER
            c2.alignment = WRAP
            row += 1

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 75
        return row + 1

    # ------------------------------------------------------------------
    # Save
    # ------------------------------------------------------------------

    def save(self, path: str | Path) -> Path:
        """Save the workbook to disk."""
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(path)
        return path
