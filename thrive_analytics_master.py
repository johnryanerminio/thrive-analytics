#!/usr/bin/env python3
"""
================================================================================
THRIVE CANNABIS - MASTER ANALYTICS SUITE v2
================================================================================
Board-ready Excel reports from Flowhub POS data.

REPORTS GENERATED:
  1. Margin_Report.xlsx         - Full Price vs Discounted analysis
  2. Deal_Performance_Report.xlsx - Deal legend, Top 10 by store
  3. Budtender_Performance_Report.xlsx - Sales Score (0-100)
  4. Customer_Insights_Report.xlsx - PERIOD-SPECIFIC metrics (not lifetime!)
  5. Rewards_Markout_Report.xlsx - Actual reward names by store

VERSION 2 CHANGES:
  - Deal Report: Added classification legend + Top 10 deals by store
  - Customer Report: Fixed to use PERIOD data, not lifetime totals
  - Customer Report: Discount calculations now working
  - Rewards Report: Shows actual reward names redeemed at each store

USAGE:
  python thrive_analytics_master_v2.py

================================================================================
"""

import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import warnings
warnings.filterwarnings('ignore')


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

THRIVE_GREEN = "2E7D32"
DARK_GREEN = "1B5E20"
LIGHT_GREEN = "E8F5E9"
HEADER_BG = "1B5E20"
ALTERNATE_ROW = "F5F5F5"
WHITE = "FFFFFF"
BLACK = "000000"
GOLD = "FFD700"
LIGHT_GOLD = "FFF8DC"
RED = "D32F2F"
LIGHT_RED = "FFEBEE"
ORANGE = "FF9800"
LIGHT_ORANGE = "FFF3E0"
TOTAL_ROW_BG = "E3F2FD"

TITLE_FONT = Font(name='Calibri', size=24, bold=True, color=DARK_GREEN)
SUBTITLE_FONT = Font(name='Calibri', size=12, italic=True, color="666666")
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color=WHITE)
DATA_FONT = Font(name='Calibri', size=10, color=BLACK)
TOTAL_FONT = Font(name='Calibri', size=10, bold=True, color=BLACK)
KPI_VALUE_FONT = Font(name='Calibri', size=28, bold=True, color=DARK_GREEN)
KPI_LABEL_FONT = Font(name='Calibri', size=10, color="666666")
SECTION_FONT = Font(name='Calibri', size=14, bold=True, color=DARK_GREEN)

HEADER_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
ALTERNATE_FILL = PatternFill(start_color=ALTERNATE_ROW, end_color=ALTERNATE_ROW, fill_type="solid")
TOTAL_FILL = PatternFill(start_color=TOTAL_ROW_BG, end_color=TOTAL_ROW_BG, fill_type="solid")
GOLD_FILL = PatternFill(start_color=LIGHT_GOLD, end_color=LIGHT_GOLD, fill_type="solid")
WARNING_FILL = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
ORANGE_FILL = PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
HEADER_BORDER = Border(
    left=Side(style='thin', color=DARK_GREEN),
    right=Side(style='thin', color=DARK_GREEN),
    top=Side(style='thin', color=DARK_GREEN),
    bottom=Side(style='medium', color=DARK_GREEN)
)
TOTAL_BORDER = Border(
    left=Side(style='thin', color='999999'),
    right=Side(style='thin', color='999999'),
    top=Side(style='medium', color='999999'),
    bottom=Side(style='medium', color='999999')
)

CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')
WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)


# =============================================================================
# FORMATTING HELPERS
# =============================================================================

def format_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = HEADER_BORDER


def format_data_cell(ws, row_num, col_num, value, col_type='text', is_total=False, highlight=None):
    cell = ws.cell(row=row_num, column=col_num)
    cell.value = value
    cell.font = TOTAL_FONT if is_total else DATA_FONT
    cell.border = TOTAL_BORDER if is_total else THIN_BORDER
    cell.alignment = RIGHT if col_type in ['currency', 'number', 'percent', 'decimal'] else LEFT
    
    if col_type == 'currency':
        cell.number_format = '"$"#,##0.00'
    elif col_type == 'percent':
        cell.number_format = '0.0"%"'
    elif col_type == 'number':
        cell.number_format = '#,##0'
    elif col_type == 'decimal':
        cell.number_format = '0.0'
    
    if highlight == 'gold':
        cell.fill = GOLD_FILL
    elif highlight == 'warning':
        cell.fill = WARNING_FILL
    elif highlight == 'orange':
        cell.fill = ORANGE_FILL
    elif is_total:
        cell.fill = TOTAL_FILL
    elif row_num % 2 == 0:
        cell.fill = ALTERNATE_FILL


def auto_column_width(ws, min_width=10, max_width=55):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def add_kpi_card(ws, row, col, value, label, format_type='currency'):
    value_cell = ws.cell(row=row, column=col)
    label_cell = ws.cell(row=row + 1, column=col)
    
    value_cell.value = value
    value_cell.font = KPI_VALUE_FONT
    value_cell.alignment = CENTER
    
    if format_type == 'currency':
        value_cell.number_format = '"$"#,##0'
    elif format_type == 'percent':
        value_cell.number_format = '0.0"%"'
    elif format_type == 'number':
        value_cell.number_format = '#,##0'
    elif format_type == 'decimal':
        value_cell.number_format = '0.0'
    
    label_cell.value = label
    label_cell.font = KPI_LABEL_FONT
    label_cell.alignment = CENTER


# =============================================================================
# DATA LOADING
# =============================================================================

def load_sales_data(filepath):
    print(f"  Loading: {filepath.name}")
    df = pd.read_csv(filepath)
    
    col_map = {
        'Pre-Discount, Pre-Tax Total': 'pre_discount_revenue',
        'Discounts': 'discounts',
        'Post-Discount, Pre-Tax Total': 'post_discount_revenue',
        'Net Profit': 'net_profit',
        'Cost': 'cost',
        'Quantity Sold': 'quantity',
        'Receipt ID': 'receipt_id',
        'Order Type': 'order_type',
        'Sold By': 'sold_by',
        'Completed At': 'completed_at',
        'Customer ID': 'customer_id',
        'Customer Name': 'customer_name',
        'Store': 'store',
        'Product': 'product',
        'Variant Type': 'category',
        'Brand': 'brand',
        'Deals Used': 'deals_used',
        'Inline/Cart Discounts Used': 'inline_discounts'
    }
    df = df.rename(columns=col_map)
    
    currency_cols = ['pre_discount_revenue', 'discounts', 'post_discount_revenue', 'net_profit', 'cost']
    for col in currency_cols:
        if col in df.columns:
            df[col] = df[col].replace(r'[\$,]', '', regex=True).astype(float)
    
    df['completed_at'] = pd.to_datetime(df['completed_at'], format='%m/%d/%Y %I:%M:%S %p', errors='coerce')
    df['sale_date'] = df['completed_at'].dt.date
    df = df.dropna(subset=['completed_at'])
    
    df['store_clean'] = df['store'].str.replace(r' - RD\d+', '', regex=True).str.strip()
    df['brand_clean'] = df['brand'].str.strip()
    df['category_clean'] = df['category'].str.strip().str.upper() if 'category' in df.columns else ''
    df['product_clean'] = df['product'].str.strip().str.upper() if 'product' in df.columns else ''
    df['deals_upper'] = df['deals_used'].fillna('').str.upper()
    df['has_discount'] = (df['discounts'] > 0)
    df['actual_revenue'] = df['post_discount_revenue']
    
    return df


def load_bt_performance(filepath):
    print(f"  Loading: {filepath.name}")
    df = pd.read_csv(filepath)
    
    currency_cols = ['Average Cart Value (pre-tax)', 'Sales (pre-tax)', 'Upsell Total Price', 'Upsell Total Profit']
    for col in currency_cols:
        if col in df.columns:
            df[col] = df[col].replace(r'[\$,]', '', regex=True).astype(float)
    
    if '% of Sales Discounted' in df.columns:
        df['% of Sales Discounted'] = df['% of Sales Discounted'].replace(r'%', '', regex=True).astype(float)
    
    return df


def load_customer_attributes(filepath):
    print(f"  Loading: {filepath.name}")
    df = pd.read_csv(filepath)
    
    col_map = {
        'ID': 'customer_id',
        'Name': 'customer_name',
        'Groups': 'groups',
        'Loyal': 'is_loyal',
        'Loyalty Points': 'loyalty_points',
    }
    df = df.rename(columns=col_map)
    return df


# =============================================================================
# CLASSIFICATION FUNCTIONS
# =============================================================================

def classify_transaction(row):
    deals = str(row.get('deals_upper', ''))
    product = str(row.get('product_clean', ''))
    actual_rev = row.get('actual_revenue', 0)
    
    if 'REWARD' in deals or 'POINT' in deals or 'REDEMPTION' in deals:
        return 'REWARD'
    if 'MARKOUT' in deals or 'MARK OUT' in deals or 'MARK-OUT' in deals:
        return 'MARKOUT'
    if 'TESTER' in product or 'TESTER' in deals:
        return 'TESTER'
    if actual_rev <= 1.00 and 'EXIT BAG' not in product:
        return 'COMP'
    return 'REGULAR'


def classify_deal_type(row):
    deals = str(row.get('deals_upper', ''))
    inline = str(row.get('inline_discounts', '')).upper() if pd.notna(row.get('inline_discounts')) else ''
    
    if not deals and not inline:
        return 'NO DEAL'
    combined = deals + ' ' + inline
    
    if any(x in combined for x in ['B1G', 'B2G', 'BOGO', '2 FOR', '3 FOR', '4 FOR', '5 FOR', '2/$', '3/$', '4/$', '5/$']):
        return 'BUNDLE'
    if '%' in combined or 'PERCENT' in combined:
        return 'PERCENT OFF'
    if any(x in combined for x in ['SENIOR', 'VETERAN', 'MILITARY', 'MEDICAL', 'INDUSTRY', 'VIP', 'EMPLOYEE']):
        return 'CUSTOMER DISCOUNT'
    if 'FOR $' in combined or 'FOR$' in combined:
        return 'PRICE DEAL'
    return 'OTHER'


def get_customer_segment(groups):
    if pd.isna(groups) or groups == '':
        return 'Regular'
    groups_upper = str(groups).upper()
    if 'INDUSTRY' in groups_upper:
        return 'Industry'
    if 'EMPLOYEE' in groups_upper:
        return 'Employee'
    if 'VETERAN' in groups_upper or 'MILITARY' in groups_upper:
        return 'Veteran'
    if 'SENIOR' in groups_upper:
        return 'Senior'
    if 'VIP' in groups_upper:
        return 'VIP'
    if 'MEDICAL' in groups_upper or 'MED' in groups_upper:
        return 'Medical'
    if 'LOCAL' in groups_upper:
        return 'Locals'
    return 'Other Group'


def extract_reward_name(deals_str):
    if pd.isna(deals_str):
        return None
    match = re.search(r'(REWARD\s*-\s*\d+\s*Points?\s*-\s*[^,]+)', str(deals_str), re.IGNORECASE)
    if match:
        return match.group(1).strip()
    if 'REWARD' in str(deals_str).upper():
        return str(deals_str).strip()
    return None


# =============================================================================
# REPORT 1: MARGIN REPORT
# =============================================================================

def create_margin_report(regular_df, output_path, date_range):
    wb = Workbook()
    
    full_price_df = regular_df[~regular_df['has_discount']]
    discounted_df = regular_df[regular_df['has_discount']]
    
    totals = {
        'total_units': regular_df['quantity'].sum(),
        'total_revenue': regular_df['actual_revenue'].sum(),
        'total_cost': regular_df['cost'].sum(),
        'net_profit': regular_df['net_profit'].sum(),
        'full_price_units': full_price_df['quantity'].sum() if len(full_price_df) > 0 else 0,
        'full_price_sales': full_price_df['actual_revenue'].sum() if len(full_price_df) > 0 else 0,
        'full_price_cost': full_price_df['cost'].sum() if len(full_price_df) > 0 else 0,
        'discounted_units': discounted_df['quantity'].sum() if len(discounted_df) > 0 else 0,
        'discounted_sales': discounted_df['actual_revenue'].sum() if len(discounted_df) > 0 else 0,
        'discounted_cost': discounted_df['cost'].sum() if len(discounted_df) > 0 else 0,
    }
    
    totals['pct_full_price'] = round(totals['full_price_sales'] / totals['total_revenue'] * 100, 1) if totals['total_revenue'] > 0 else 0
    totals['pct_discounted'] = round(totals['discounted_sales'] / totals['total_revenue'] * 100, 1) if totals['total_revenue'] > 0 else 0
    totals['full_price_margin'] = round((totals['full_price_sales'] - totals['full_price_cost']) / totals['full_price_sales'] * 100, 1) if totals['full_price_sales'] > 0 else 0
    totals['discounted_margin'] = round((totals['discounted_sales'] - totals['discounted_cost']) / totals['discounted_sales'] * 100, 1) if totals['discounted_sales'] > 0 else 0
    totals['blended_margin'] = round((totals['total_revenue'] - totals['total_cost']) / totals['total_revenue'] * 100, 1) if totals['total_revenue'] > 0 else 0
    
    # Executive Summary
    ws = wb.active
    ws.title = "Executive Summary"
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    ws.cell(row=1, column=1).value = "THRIVE CANNABIS"
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    
    ws.cell(row=2, column=1).value = f"Margin Performance Report  |  {date_range}  |  Generated {datetime.now().strftime('%B %d, %Y')}"
    ws.cell(row=2, column=1).font = SUBTITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    
    ws.cell(row=5, column=1).value = "REVENUE OVERVIEW"
    ws.cell(row=5, column=1).font = SECTION_FONT
    add_kpi_card(ws, 7, 1, totals['total_revenue'], "TOTAL REVENUE", 'currency')
    add_kpi_card(ws, 7, 3, totals['full_price_sales'], "FULL PRICE SALES", 'currency')
    add_kpi_card(ws, 7, 5, totals['discounted_sales'], "DISCOUNTED SALES", 'currency')
    add_kpi_card(ws, 7, 7, totals['net_profit'], "NET PROFIT", 'currency')
    
    ws.cell(row=11, column=1).value = "SALES MIX"
    ws.cell(row=11, column=1).font = SECTION_FONT
    add_kpi_card(ws, 13, 1, totals['pct_full_price'], "% AT FULL PRICE", 'percent')
    add_kpi_card(ws, 13, 3, totals['pct_discounted'], "% ON DISCOUNT", 'percent')
    add_kpi_card(ws, 13, 5, totals['total_units'], "TOTAL UNITS", 'number')
    
    ws.cell(row=17, column=1).value = "MARGIN ANALYSIS"
    ws.cell(row=17, column=1).font = SECTION_FONT
    add_kpi_card(ws, 19, 1, totals['full_price_margin'], "FULL PRICE MARGIN", 'percent')
    add_kpi_card(ws, 19, 3, totals['discounted_margin'], "DISCOUNTED MARGIN", 'percent')
    add_kpi_card(ws, 19, 5, totals['blended_margin'], "BLENDED MARGIN", 'percent')
    add_kpi_card(ws, 19, 7, totals['full_price_margin'] - totals['discounted_margin'], "MARGIN GAP (pts)", 'decimal')
    
    # Data sheets helper
    def calc_metrics(df, group_col):
        fp = df[~df['has_discount']]
        disc = df[df['has_discount']]
        
        fp_agg = fp.groupby(group_col).agg(
            full_price_units=('quantity', 'sum'),
            full_price_sales=('actual_revenue', 'sum'),
            full_price_cost=('cost', 'sum'),
        ).reset_index() if len(fp) > 0 else pd.DataFrame(columns=[group_col, 'full_price_units', 'full_price_sales', 'full_price_cost'])
        
        disc_agg = disc.groupby(group_col).agg(
            discounted_units=('quantity', 'sum'),
            discounted_sales=('actual_revenue', 'sum'),
            discounted_cost=('cost', 'sum'),
        ).reset_index() if len(disc) > 0 else pd.DataFrame(columns=[group_col, 'discounted_units', 'discounted_sales', 'discounted_cost'])
        
        total_agg = df.groupby(group_col).agg(
            total_units=('quantity', 'sum'),
            total_revenue=('actual_revenue', 'sum'),
            total_cost=('cost', 'sum'),
            net_profit=('net_profit', 'sum'),
        ).reset_index()
        
        result = total_agg.merge(fp_agg, on=group_col, how='left').merge(disc_agg, on=group_col, how='left')
        for col in ['full_price_units', 'full_price_sales', 'full_price_cost', 'discounted_units', 'discounted_sales', 'discounted_cost']:
            if col in result.columns:
                result[col] = result[col].fillna(0)
        
        result['pct_full_price'] = (result['full_price_sales'] / result['total_revenue'].replace(0, np.nan) * 100).round(1)
        result['pct_discounted'] = (result['discounted_sales'] / result['total_revenue'].replace(0, np.nan) * 100).round(1)
        result['full_price_margin'] = ((result['full_price_sales'] - result['full_price_cost']) / result['full_price_sales'].replace(0, np.nan) * 100).round(1)
        result['discounted_margin'] = ((result['discounted_sales'] - result['discounted_cost']) / result['discounted_sales'].replace(0, np.nan) * 100).round(1)
        result['blended_margin'] = ((result['total_revenue'] - result['total_cost']) / result['total_revenue'].replace(0, np.nan) * 100).round(1)
        
        result = result.rename(columns={group_col: 'name'})
        return result.sort_values('total_revenue', ascending=False)
    
    COLS = [
        ('name', 'text', 'Name'),
        ('full_price_units', 'number', 'Full Price Units'),
        ('discounted_units', 'number', 'Discounted Units'),
        ('total_units', 'number', 'Total Units'),
        ('full_price_sales', 'currency', 'Full Price Sales'),
        ('discounted_sales', 'currency', 'Discounted Sales'),
        ('total_revenue', 'currency', 'Total Revenue'),
        ('pct_full_price', 'percent', '% Full Price'),
        ('pct_discounted', 'percent', '% Discounted'),
        ('full_price_margin', 'percent', 'FP Margin'),
        ('discounted_margin', 'percent', 'Disc Margin'),
        ('blended_margin', 'percent', 'Blended Margin'),
        ('net_profit', 'currency', 'Net Profit'),
    ]
    
    def write_sheet(wb, data_df, sheet_name):
        ws = wb.create_sheet(title=sheet_name)
        for col_num, (_, _, label) in enumerate(COLS, 1):
            ws.cell(row=1, column=col_num).value = label
        format_header_row(ws, 1, len(COLS))
        
        row = 2
        for _, row_data in data_df.iterrows():
            for col_num, (col_key, col_type, _) in enumerate(COLS, 1):
                val = row_data.get(col_key, 0)
                format_data_cell(ws, row, col_num, 0 if pd.isna(val) else val, col_type)
            row += 1
        
        if len(data_df) > 1:
            format_data_cell(ws, row, 1, 'TOTAL', 'text', is_total=True)
            format_data_cell(ws, row, 2, data_df['full_price_units'].sum(), 'number', is_total=True)
            format_data_cell(ws, row, 3, data_df['discounted_units'].sum(), 'number', is_total=True)
            format_data_cell(ws, row, 4, data_df['total_units'].sum(), 'number', is_total=True)
            format_data_cell(ws, row, 5, data_df['full_price_sales'].sum(), 'currency', is_total=True)
            format_data_cell(ws, row, 6, data_df['discounted_sales'].sum(), 'currency', is_total=True)
            format_data_cell(ws, row, 7, data_df['total_revenue'].sum(), 'currency', is_total=True)
            tr = data_df['total_revenue'].sum()
            fps = data_df['full_price_sales'].sum()
            tc = data_df['total_cost'].sum()
            format_data_cell(ws, row, 8, fps / tr * 100 if tr > 0 else 0, 'percent', is_total=True)
            format_data_cell(ws, row, 9, (tr - fps) / tr * 100 if tr > 0 else 0, 'percent', is_total=True)
            format_data_cell(ws, row, 10, totals['full_price_margin'], 'percent', is_total=True)
            format_data_cell(ws, row, 11, totals['discounted_margin'], 'percent', is_total=True)
            format_data_cell(ws, row, 12, (tr - tc) / tr * 100 if tr > 0 else 0, 'percent', is_total=True)
            format_data_cell(ws, row, 13, data_df['net_profit'].sum(), 'currency', is_total=True)
        
        auto_column_width(ws)
        ws.freeze_panes = 'A2'
    
    write_sheet(wb, calc_metrics(regular_df, 'store_clean'), "By Store")
    write_sheet(wb, calc_metrics(regular_df, 'brand_clean'), "By Brand")
    write_sheet(wb, calc_metrics(regular_df, 'category_clean'), "By Category")
    write_sheet(wb, calc_metrics(regular_df, 'deal_type'), "By Deal Type")
    
    wb.save(output_path)
    return totals


# =============================================================================
# REPORT 2: DEAL PERFORMANCE (with Legend + Top 10 by Store)
# =============================================================================

def create_deal_report(regular_df, output_path, date_range):
    wb = Workbook()
    ws = wb.active
    ws.title = "Deal Summary"
    
    # Title
    ws.cell(row=1, column=1).value = "THRIVE CANNABIS"
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    
    ws.cell(row=2, column=1).value = f"Deal Performance Report  |  {date_range}"
    ws.cell(row=2, column=1).font = SUBTITLE_FONT
    
    # LEGEND
    ws.cell(row=4, column=1).value = "DEAL CLASSIFICATION KEY"
    ws.cell(row=4, column=1).font = SECTION_FONT
    
    legend = [
        ("NO DEAL", "Items sold at full price - no discount applied"),
        ("BUNDLE", "Multi-buy deals: BOGO, B1G1, 2 FOR $X, 3/$X, 4/$X, 5/$X, etc."),
        ("PERCENT OFF", "Percentage discounts: 10% OFF, 20% OFF, 25% OFF, etc."),
        ("CUSTOMER DISCOUNT", "Customer-type discounts: Senior, Veteran, Military, Industry, Medical, VIP, Employee"),
        ("PRICE DEAL", "Fixed price deals: 2 FOR $25, Eighths FOR $X, etc."),
        ("OTHER", "All other promotional deals not matching above categories"),
    ]
    
    ws.cell(row=5, column=1).value = "Category"
    ws.cell(row=5, column=2).value = "What It Includes"
    format_header_row(ws, 5, 2)
    
    for row_num, (cat, desc) in enumerate(legend, 6):
        c1 = ws.cell(row=row_num, column=1)
        c1.value = cat
        c1.font = Font(name='Calibri', size=10, bold=True)
        c1.fill = LIGHT_GREEN_FILL
        c1.border = THIN_BORDER
        
        c2 = ws.cell(row=row_num, column=2)
        c2.value = desc
        c2.font = DATA_FONT
        c2.fill = LIGHT_GREEN_FILL
        c2.border = THIN_BORDER
        c2.alignment = WRAP
    
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 75
    
    # Deal Type Summary
    ws.cell(row=13, column=1).value = "PERFORMANCE BY DEAL TYPE"
    ws.cell(row=13, column=1).font = SECTION_FONT
    
    deal_summary = regular_df.groupby('deal_type').agg(
        transactions=('receipt_id', 'nunique'),
        units=('quantity', 'sum'),
        full_price_revenue=('pre_discount_revenue', 'sum'),
        actual_revenue=('actual_revenue', 'sum'),
        discounts=('discounts', 'sum'),
        cost=('cost', 'sum'),
        net_profit=('net_profit', 'sum'),
    ).reset_index()
    deal_summary['discount_rate'] = (deal_summary['discounts'] / deal_summary['full_price_revenue'] * 100).round(1)
    deal_summary['margin'] = ((deal_summary['actual_revenue'] - deal_summary['cost']) / deal_summary['actual_revenue'] * 100).round(1)
    deal_summary = deal_summary.sort_values('actual_revenue', ascending=False)
    
    cols = [
        ('deal_type', 'text', 'Deal Type'),
        ('transactions', 'number', 'Transactions'),
        ('units', 'number', 'Units'),
        ('full_price_revenue', 'currency', 'Full Price Revenue'),
        ('discounts', 'currency', 'Discounts Given'),
        ('actual_revenue', 'currency', 'Actual Revenue'),
        ('discount_rate', 'percent', 'Discount Rate'),
        ('margin', 'percent', 'Margin'),
        ('net_profit', 'currency', 'Net Profit'),
    ]
    
    row = 14
    for col_num, (_, _, label) in enumerate(cols, 1):
        ws.cell(row=row, column=col_num).value = label
    format_header_row(ws, row, len(cols))
    
    row += 1
    for _, data in deal_summary.iterrows():
        for col_num, (col_key, col_type, _) in enumerate(cols, 1):
            format_data_cell(ws, row, col_num, data.get(col_key, 0), col_type)
        row += 1
    
    # Total row
    row += 1
    format_data_cell(ws, row, 1, 'TOTAL', 'text', is_total=True)
    format_data_cell(ws, row, 2, deal_summary['transactions'].sum(), 'number', is_total=True)
    format_data_cell(ws, row, 3, deal_summary['units'].sum(), 'number', is_total=True)
    format_data_cell(ws, row, 4, deal_summary['full_price_revenue'].sum(), 'currency', is_total=True)
    format_data_cell(ws, row, 5, deal_summary['discounts'].sum(), 'currency', is_total=True)
    format_data_cell(ws, row, 6, deal_summary['actual_revenue'].sum(), 'currency', is_total=True)
    tfp = deal_summary['full_price_revenue'].sum()
    td = deal_summary['discounts'].sum()
    tr = deal_summary['actual_revenue'].sum()
    tc = deal_summary['cost'].sum()
    format_data_cell(ws, row, 7, td / tfp * 100 if tfp > 0 else 0, 'percent', is_total=True)
    format_data_cell(ws, row, 8, (tr - tc) / tr * 100 if tr > 0 else 0, 'percent', is_total=True)
    format_data_cell(ws, row, 9, deal_summary['net_profit'].sum(), 'currency', is_total=True)
    
    for col in range(3, 10):
        ws.column_dimensions[get_column_letter(col)].width = 16
    
    # Extract individual deals
    def extract_deals(deals_str):
        if pd.isna(deals_str) or deals_str == '':
            return []
        return [d.strip() for d in str(deals_str).split(',') if d.strip()]
    
    deals_expanded = []
    for _, row_data in regular_df.iterrows():
        deals = extract_deals(row_data['deals_used'])
        if deals:
            for deal in deals:
                deals_expanded.append({
                    'deal_name': deal,
                    'receipt_id': row_data['receipt_id'],  # Track receipt for unique counting
                    'store': row_data['store_clean'],
                    'revenue': row_data['actual_revenue'],
                    'discounts': row_data['discounts'],
                    'quantity': row_data['quantity'],
                    'cost': row_data['cost'],
                })
    
    # Top 50 Deals sheet
    if deals_expanded:
        deals_df = pd.DataFrame(deals_expanded)
        
        ws2 = wb.create_sheet(title="Top 50 Deals")
        top_deals = deals_df.groupby('deal_name').agg(
            times_used=('receipt_id', 'nunique'),  # Count unique transactions, not line items!
            units=('quantity', 'sum'),
            revenue=('revenue', 'sum'),
            discounts=('discounts', 'sum'),
            cost=('cost', 'sum'),
        ).reset_index()
        top_deals['margin'] = ((top_deals['revenue'] - top_deals['cost']) / top_deals['revenue'] * 100).round(1)
        top_deals = top_deals.sort_values('times_used', ascending=False).head(50)
        
        cols = [
            ('deal_name', 'text', 'Deal Name'),
            ('times_used', 'number', 'Times Used'),
            ('units', 'number', 'Units'),
            ('revenue', 'currency', 'Revenue'),
            ('discounts', 'currency', 'Discounts'),
            ('margin', 'percent', 'Margin'),
        ]
        
        for col_num, (_, _, label) in enumerate(cols, 1):
            ws2.cell(row=1, column=col_num).value = label
        format_header_row(ws2, 1, len(cols))
        
        for row_num, (_, data) in enumerate(top_deals.iterrows(), 2):
            for col_num, (col_key, col_type, _) in enumerate(cols, 1):
                hl = 'gold' if row_num <= 11 else None
                format_data_cell(ws2, row_num, col_num, data.get(col_key, 0), col_type, highlight=hl)
        
        auto_column_width(ws2)
        ws2.freeze_panes = 'A2'
        
        # Top 10 by Store tabs
        stores = sorted(deals_df['store'].dropna().unique())
        for store in stores:
            store_deals = deals_df[deals_df['store'] == store]
            top_store = store_deals.groupby('deal_name').agg(
                times_used=('receipt_id', 'nunique'),  # Count unique transactions, not line items!
                units=('quantity', 'sum'),
                revenue=('revenue', 'sum'),
                discounts=('discounts', 'sum'),
                cost=('cost', 'sum'),
            ).reset_index()
            top_store['margin'] = ((top_store['revenue'] - top_store['cost']) / top_store['revenue'] * 100).round(1)
            top_store = top_store.sort_values('times_used', ascending=False).head(10)
            
            if len(top_store) == 0:
                continue
            
            short_name = store.replace('Thrive ', '').replace('Cannabis ', '')[:12]
            ws_store = wb.create_sheet(title=f"Top 10 - {short_name}")
            
            for col_num, (_, _, label) in enumerate(cols, 1):
                ws_store.cell(row=1, column=col_num).value = label
            format_header_row(ws_store, 1, len(cols))
            
            for row_num, (_, data) in enumerate(top_store.iterrows(), 2):
                for col_num, (col_key, col_type, _) in enumerate(cols, 1):
                    hl = 'gold' if row_num <= 4 else None
                    format_data_cell(ws_store, row_num, col_num, data.get(col_key, 0), col_type, highlight=hl)
            
            auto_column_width(ws_store)
            ws_store.freeze_panes = 'A2'
    
    wb.save(output_path)


# =============================================================================
# REPORT 3: BUDTENDER PERFORMANCE
# =============================================================================

def create_budtender_report(bt_df, sales_df, output_path, date_range):
    wb = Workbook()
    
    bt_df = bt_df.rename(columns={
        'Name': 'budtender', 'Store': 'store',
        'Average Cart Value (pre-tax)': 'avg_cart_value',
        'Total Units Sold': 'units_sold',
        'Average Units Per Cart': 'avg_units_per_cart',
        'Number of Carts': 'num_transactions',
        'Sales (pre-tax)': 'total_sales',
        '% of Sales Discounted': 'pct_sales_discounted',
        'Customers Enrolled In Loyalty': 'loyalty_enrollments',
    })
    
    bt_df['store_clean'] = bt_df['store'].str.replace(r' - RD\d+', '', regex=True).str.strip()
    
    # Face-to-face %
    if 'order_type' in sales_df.columns:
        f2f = sales_df.groupby('sold_by').apply(
            lambda x: (x['order_type'].str.upper().str.contains('WALK|IN-STORE|FACE', na=False).sum() / len(x) * 100) if len(x) > 0 else 0
        ).reset_index()
        f2f.columns = ['budtender', 'face_to_face_pct']
        bt_df = bt_df.merge(f2f, on='budtender', how='left')
        bt_df['face_to_face_pct'] = bt_df['face_to_face_pct'].fillna(0)
    else:
        bt_df['face_to_face_pct'] = 0
    
    # Sales Score
    bt_min = bt_df[bt_df['num_transactions'] >= 5].copy()
    if len(bt_min) > 0:
        bt_min['cart_score'] = ((bt_min['avg_cart_value'] - bt_min['avg_cart_value'].min()) / 
                                (bt_min['avg_cart_value'].max() - bt_min['avg_cart_value'].min() + 0.01) * 30)
        bt_min['units_score'] = ((bt_min['avg_units_per_cart'] - bt_min['avg_units_per_cart'].min()) / 
                                 (bt_min['avg_units_per_cart'].max() - bt_min['avg_units_per_cart'].min() + 0.01) * 25)
        bt_min['discount_score'] = ((100 - bt_min['pct_sales_discounted']) / 100 * 20)
        bt_min['loyalty_score'] = (bt_min['loyalty_enrollments'] / bt_min['loyalty_enrollments'].max() * 15) if bt_min['loyalty_enrollments'].max() > 0 else 0
        bt_min['f2f_score'] = (bt_min['face_to_face_pct'] / 100 * 10)
        bt_min['sales_score'] = (bt_min['cart_score'] + bt_min['units_score'] + bt_min['discount_score'] + 
                                  bt_min['loyalty_score'] + bt_min['f2f_score']).round(0)
    else:
        bt_min = bt_df.copy()
        bt_min['sales_score'] = 0
    
    bt_min = bt_min.sort_values('sales_score', ascending=False)
    
    def get_tier(score):
        if score >= 70: return 'Top Performer'
        elif score >= 50: return 'Solid'
        elif score >= 30: return 'Developing'
        else: return 'Needs Coaching'
    
    bt_min['tier'] = bt_min['sales_score'].apply(get_tier)
    
    # Executive Summary
    ws = wb.active
    ws.title = "Executive Summary"
    
    ws.cell(row=1, column=1).value = "THRIVE CANNABIS"
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    
    ws.cell(row=2, column=1).value = f"Budtender Performance Report  |  {date_range}"
    ws.cell(row=2, column=1).font = SUBTITLE_FONT
    
    ws.cell(row=4, column=1).value = "TEAM OVERVIEW"
    ws.cell(row=4, column=1).font = SECTION_FONT
    
    add_kpi_card(ws, 6, 1, len(bt_min), "TOTAL BUDTENDERS", 'number')
    add_kpi_card(ws, 6, 3, bt_min['sales_score'].mean(), "AVG SALES SCORE", 'number')
    add_kpi_card(ws, 6, 5, len(bt_min[bt_min['tier'] == 'Top Performer']), "TOP PERFORMERS", 'number')
    add_kpi_card(ws, 6, 7, len(bt_min[bt_min['tier'] == 'Needs Coaching']), "NEEDS COACHING", 'number')
    
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # All Rankings
    ws2 = wb.create_sheet(title="All Rankings")
    
    cols = [
        ('budtender', 'text', 'Budtender'),
        ('store_clean', 'text', 'Store'),
        ('sales_score', 'number', 'Sales Score'),
        ('tier', 'text', 'Tier'),
        ('num_transactions', 'number', 'Transactions'),
        ('total_sales', 'currency', 'Total Sales'),
        ('avg_cart_value', 'currency', 'Avg Cart'),
        ('avg_units_per_cart', 'decimal', 'Units/Cart'),
        ('pct_sales_discounted', 'percent', 'Discount %'),
        ('face_to_face_pct', 'percent', 'F2F %'),
    ]
    
    for col_num, (_, _, label) in enumerate(cols, 1):
        ws2.cell(row=1, column=col_num).value = label
    format_header_row(ws2, 1, len(cols))
    
    for row_num, (_, data) in enumerate(bt_min.iterrows(), 2):
        for col_num, (col_key, col_type, _) in enumerate(cols, 1):
            hl = 'gold' if data.get('tier') == 'Top Performer' else ('warning' if data.get('tier') == 'Needs Coaching' else None)
            format_data_cell(ws2, row_num, col_num, data.get(col_key, 0), col_type, highlight=hl)
    
    auto_column_width(ws2)
    ws2.freeze_panes = 'A2'
    
    # Store tabs
    stores = bt_min['store_clean'].dropna().unique()
    for store in sorted([s for s in stores if pd.notna(s)]):
        store_bt = bt_min[bt_min['store_clean'] == store]
        if len(store_bt) == 0:
            continue
        
        short_name = store.replace('Thrive ', '').replace('Cannabis ', '')[:20]
        ws_store = wb.create_sheet(title=f"{short_name}")
        
        for col_num, (_, _, label) in enumerate(cols, 1):
            ws_store.cell(row=1, column=col_num).value = label
        format_header_row(ws_store, 1, len(cols))
        
        for row_num, (_, data) in enumerate(store_bt.iterrows(), 2):
            for col_num, (col_key, col_type, _) in enumerate(cols, 1):
                hl = 'gold' if row_num <= 4 else None
                format_data_cell(ws_store, row_num, col_num, data.get(col_key, 0), col_type, highlight=hl)
        
        auto_column_width(ws_store)
        ws_store.freeze_panes = 'A2'
    
    wb.save(output_path)


# =============================================================================
# REPORT 4: CUSTOMER INSIGHTS (Period-Specific!)
# =============================================================================

def create_customer_report(sales_df, cust_attr_df, output_path, date_range):
    wb = Workbook()
    
    # Filter regular sales
    def is_regular(row):
        deals = str(row.get('deals_upper', ''))
        rev = row.get('actual_revenue', 0)
        if 'REWARD' in deals or 'POINT' in deals or 'MARK' in deals:
            return False
        if rev <= 1.00:
            return False
        return True
    
    sales_df['is_regular'] = sales_df.apply(is_regular, axis=1)
    regular = sales_df[sales_df['is_regular']].copy()
    
    # Calculate period metrics from SALES data
    cust_metrics = regular.groupby('customer_id').agg(
        customer_name=('customer_name', 'first'),
        transactions=('receipt_id', 'nunique'),
        total_spent=('actual_revenue', 'sum'),
        total_discounts=('discounts', 'sum'),
        total_units=('quantity', 'sum'),
        primary_store=('store_clean', lambda x: x.value_counts().index[0] if len(x) > 0 else 'Unknown'),
    ).reset_index()
    
    # Avg transaction
    trans_totals = regular.groupby(['customer_id', 'receipt_id'])['actual_revenue'].sum().reset_index()
    avg_by_cust = trans_totals.groupby('customer_id')['actual_revenue'].mean().reset_index()
    avg_by_cust.columns = ['customer_id', 'avg_transaction']
    cust_metrics = cust_metrics.merge(avg_by_cust, on='customer_id', how='left')
    
    # Discount rate
    cust_metrics['discount_rate'] = (cust_metrics['total_discounts'] / 
        (cust_metrics['total_spent'] + cust_metrics['total_discounts']).replace(0, np.nan) * 100).round(1)
    
    # Merge attributes
    if cust_attr_df is not None:
        cust_metrics = cust_metrics.merge(
            cust_attr_df[['customer_id', 'groups', 'is_loyal']], 
            on='customer_id', how='left'
        )
    else:
        cust_metrics['groups'] = ''
        cust_metrics['is_loyal'] = 'No'
    
    cust_metrics['segment'] = cust_metrics['groups'].apply(get_customer_segment)
    
    # Totals
    total_rev = regular['actual_revenue'].sum()
    total_disc = regular['discounts'].sum()
    total_trans = regular['receipt_id'].nunique()
    total_cust = cust_metrics['customer_id'].nunique()
    loyalty_cust = (cust_metrics['is_loyal'] == 'Yes').sum()
    loyalty_rate = loyalty_cust / total_cust * 100 if total_cust > 0 else 0
    disc_rate = total_disc / (total_rev + total_disc) * 100 if (total_rev + total_disc) > 0 else 0
    
    # Executive Summary
    ws = wb.active
    ws.title = "Executive Summary"
    
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    ws.cell(row=1, column=1).value = "THRIVE CANNABIS"
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    
    ws.cell(row=2, column=1).value = f"Customer Insights Report  |  {date_range}  |  Generated {datetime.now().strftime('%B %d, %Y')}"
    ws.cell(row=2, column=1).font = SUBTITLE_FONT
    
    ws.cell(row=3, column=1).value = f"📅 DATA PERIOD: {date_range}"
    ws.cell(row=3, column=1).font = Font(name='Calibri', size=11, bold=True, color="D32F2F")
    
    ws.cell(row=5, column=1).value = "CUSTOMER OVERVIEW"
    ws.cell(row=5, column=1).font = SECTION_FONT
    
    add_kpi_card(ws, 7, 1, total_cust, "UNIQUE CUSTOMERS", 'number')
    add_kpi_card(ws, 7, 3, total_rev, "TOTAL REVENUE", 'currency')
    add_kpi_card(ws, 7, 5, total_rev / total_cust if total_cust > 0 else 0, "REVENUE/CUSTOMER", 'currency')
    add_kpi_card(ws, 7, 7, loyalty_rate, "LOYALTY RATE", 'percent')
    
    ws.cell(row=11, column=1).value = "TRANSACTION METRICS"
    ws.cell(row=11, column=1).font = SECTION_FONT
    
    add_kpi_card(ws, 13, 1, total_trans, "TOTAL TRANSACTIONS", 'number')
    add_kpi_card(ws, 13, 3, total_rev / total_trans if total_trans > 0 else 0, "AVG TRANSACTION", 'currency')
    add_kpi_card(ws, 13, 5, total_disc, "TOTAL DISCOUNTS", 'currency')
    add_kpi_card(ws, 13, 7, disc_rate, "DISCOUNT RATE", 'percent')
    
    # Segments
    ws2 = wb.create_sheet(title="Customer Segments")
    
    seg_sum = cust_metrics.groupby('segment').agg(
        customers=('customer_id', 'count'),
        total_revenue=('total_spent', 'sum'),
        total_discounts=('total_discounts', 'sum'),
    ).reset_index()
    seg_sum['rev_per_cust'] = (seg_sum['total_revenue'] / seg_sum['customers']).round(2)
    seg_sum['discount_rate'] = (seg_sum['total_discounts'] / (seg_sum['total_revenue'] + seg_sum['total_discounts']).replace(0, np.nan) * 100).round(1)
    seg_sum['pct_of_cust'] = (seg_sum['customers'] / total_cust * 100).round(1)
    seg_sum['pct_of_rev'] = (seg_sum['total_revenue'] / total_rev * 100).round(1)
    seg_sum = seg_sum.sort_values('total_revenue', ascending=False)
    
    cols = [
        ('segment', 'text', 'Segment'),
        ('customers', 'number', 'Customers'),
        ('pct_of_cust', 'percent', '% of Customers'),
        ('total_revenue', 'currency', 'Total Revenue'),
        ('pct_of_rev', 'percent', '% of Revenue'),
        ('rev_per_cust', 'currency', 'Rev/Customer'),
        ('total_discounts', 'currency', 'Total Discounts'),
        ('discount_rate', 'percent', 'Discount Rate'),
    ]
    
    for col_num, (_, _, label) in enumerate(cols, 1):
        ws2.cell(row=1, column=col_num).value = label
    format_header_row(ws2, 1, len(cols))
    
    for row_num, (_, data) in enumerate(seg_sum.iterrows(), 2):
        for col_num, (col_key, col_type, _) in enumerate(cols, 1):
            val = data.get(col_key, 0)
            format_data_cell(ws2, row_num, col_num, 0 if pd.isna(val) else val, col_type)
    
    auto_column_width(ws2)
    ws2.freeze_panes = 'A2'
    
    # High Value Customers
    ws3 = wb.create_sheet(title="High Value Customers")
    
    ws3.cell(row=1, column=1).value = f"Top 50 Customers by Spend During {date_range}"
    ws3.cell(row=1, column=1).font = SECTION_FONT
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    
    top_cust = cust_metrics.nlargest(50, 'total_spent').copy()
    top_cust['rank'] = range(1, len(top_cust) + 1)
    
    cols = [
        ('rank', 'number', 'Rank'),
        ('customer_name', 'text', 'Customer Name'),
        ('primary_store', 'text', 'Primary Store'),
        ('segment', 'text', 'Segment'),
        ('is_loyal', 'text', 'Loyal'),
        ('total_spent', 'currency', 'Period Spend'),
        ('transactions', 'number', 'Transactions'),
        ('avg_transaction', 'currency', 'Avg Transaction'),
        ('total_discounts', 'currency', 'Discounts'),
        ('discount_rate', 'percent', 'Discount Rate'),
    ]
    
    for col_num, (_, _, label) in enumerate(cols, 1):
        ws3.cell(row=3, column=col_num).value = label
    format_header_row(ws3, 3, len(cols))
    
    for row_num, (_, data) in enumerate(top_cust.iterrows(), 4):
        for col_num, (col_key, col_type, _) in enumerate(cols, 1):
            val = data.get(col_key, 0)
            hl = 'gold' if row_num <= 13 else None
            format_data_cell(ws3, row_num, col_num, 0 if pd.isna(val) else val, col_type, highlight=hl)
    
    auto_column_width(ws3)
    ws3.freeze_panes = 'A4'
    
    # Top 50 by Store
    stores = sorted(cust_metrics['primary_store'].dropna().unique())
    for store in stores:
        if pd.isna(store) or store == 'Unknown':
            continue
        
        store_cust = cust_metrics[cust_metrics['primary_store'] == store].nlargest(50, 'total_spent').copy()
        if len(store_cust) == 0:
            continue
        
        store_cust['rank'] = range(1, len(store_cust) + 1)
        
        short_name = store.replace('Thrive ', '').replace('Cannabis ', '')[:12]
        ws_store = wb.create_sheet(title=f"Top 50 - {short_name}")
        
        ws_store.cell(row=1, column=1).value = f"{store} - Top 50 ({date_range})"
        ws_store.cell(row=1, column=1).font = SECTION_FONT
        
        for col_num, (_, _, label) in enumerate(cols, 1):
            ws_store.cell(row=3, column=col_num).value = label
        format_header_row(ws_store, 3, len(cols))
        
        for row_num, (_, data) in enumerate(store_cust.iterrows(), 4):
            for col_num, (col_key, col_type, _) in enumerate(cols, 1):
                val = data.get(col_key, 0)
                hl = 'gold' if row_num <= 13 else None
                format_data_cell(ws_store, row_num, col_num, 0 if pd.isna(val) else val, col_type, highlight=hl)
        
        auto_column_width(ws_store)
        ws_store.freeze_panes = 'A4'
    
    wb.save(output_path)


# =============================================================================
# REPORT 5: REWARDS & MARKOUT (with actual reward names by store)
# =============================================================================

def create_rewards_report(sales_df, output_path, date_range):
    wb = Workbook()
    
    rewards_df = sales_df[sales_df['deals_upper'].str.contains('REWARD|POINT|REDEMPTION', na=False)].copy()
    rewards_df['reward_name'] = rewards_df['deals_used'].apply(extract_reward_name)
    
    markouts_df = sales_df[sales_df['deals_upper'].str.contains('MARK', na=False)].copy()
    
    days = max((sales_df['sale_date'].max() - sales_df['sale_date'].min()).days + 1, 1)
    
    rewards_cost = rewards_df['cost'].sum()
    rewards_collected = rewards_df['actual_revenue'].sum()
    rewards_net = rewards_cost - rewards_collected
    
    markouts_cost = markouts_df['cost'].sum()
    markouts_collected = markouts_df['actual_revenue'].sum()
    markouts_net = markouts_cost - markouts_collected
    
    total_net = rewards_net + markouts_net
    monthly = total_net / days * 30
    
    # Executive Summary
    ws = wb.active
    ws.title = "Executive Summary"
    
    ws.cell(row=1, column=1).value = "THRIVE CANNABIS"
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    
    ws.cell(row=2, column=1).value = f"Rewards & Markout Report  |  {date_range}"
    ws.cell(row=2, column=1).font = SUBTITLE_FONT
    
    ws.cell(row=4, column=1).value = "PROGRAM COSTS"
    ws.cell(row=4, column=1).font = SECTION_FONT
    
    add_kpi_card(ws, 6, 1, rewards_net, "REWARDS NET COST", 'currency')
    add_kpi_card(ws, 6, 3, markouts_net, "MARKOUTS NET COST", 'currency')
    add_kpi_card(ws, 6, 5, total_net, "TOTAL NET COST", 'currency')
    add_kpi_card(ws, 6, 7, monthly, "MONTHLY PROJECTION", 'currency')
    
    ws.cell(row=10, column=1).value = "USAGE STATS"
    ws.cell(row=10, column=1).font = SECTION_FONT
    
    add_kpi_card(ws, 12, 1, len(rewards_df), "REWARD REDEMPTIONS", 'number')
    add_kpi_card(ws, 12, 3, rewards_df['customer_id'].nunique() if len(rewards_df) > 0 else 0, "UNIQUE CUSTOMERS", 'number')
    add_kpi_card(ws, 12, 5, len(markouts_df), "MARKOUT TRANSACTIONS", 'number')
    add_kpi_card(ws, 12, 7, markouts_df['customer_name'].nunique() if len(markouts_df) > 0 else 0, "EMPLOYEES USING", 'number')
    
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # All Rewards
    if len(rewards_df) > 0:
        ws2 = wb.create_sheet(title="All Rewards")
        
        reward_sum = rewards_df.groupby('reward_name').agg(
            redemptions=('receipt_id', 'count'),
            units=('quantity', 'sum'),
            retail_value=('pre_discount_revenue', 'sum'),
            cost=('cost', 'sum'),
            collected=('actual_revenue', 'sum'),
        ).reset_index()
        reward_sum['net_cost'] = reward_sum['cost'] - reward_sum['collected']
        reward_sum['pct'] = (reward_sum['net_cost'] / rewards_net * 100).round(1)
        reward_sum = reward_sum.sort_values('net_cost', ascending=False)
        
        cols = [
            ('reward_name', 'text', 'Reward Name'),
            ('redemptions', 'number', 'Redemptions'),
            ('units', 'number', 'Units'),
            ('retail_value', 'currency', 'Retail Value'),
            ('cost', 'currency', 'Product Cost'),
            ('net_cost', 'currency', 'Net Cost'),
            ('pct', 'percent', '% of Total'),
        ]
        
        for col_num, (_, _, label) in enumerate(cols, 1):
            ws2.cell(row=1, column=col_num).value = label
        format_header_row(ws2, 1, len(cols))
        
        for row_num, (_, data) in enumerate(reward_sum.iterrows(), 2):
            for col_num, (col_key, col_type, _) in enumerate(cols, 1):
                hl = 'warning' if row_num <= 4 else None
                format_data_cell(ws2, row_num, col_num, data.get(col_key, 0), col_type, highlight=hl)
        
        auto_column_width(ws2)
        ws2.freeze_panes = 'A2'
        
        # Rewards by Store tabs
        stores = sorted(rewards_df['store_clean'].dropna().unique())
        for store in stores:
            store_rewards = rewards_df[rewards_df['store_clean'] == store]
            
            store_sum = store_rewards.groupby('reward_name').agg(
                redemptions=('receipt_id', 'count'),
                units=('quantity', 'sum'),
                retail_value=('pre_discount_revenue', 'sum'),
                cost=('cost', 'sum'),
                collected=('actual_revenue', 'sum'),
            ).reset_index()
            store_sum['net_cost'] = store_sum['cost'] - store_sum['collected']
            store_sum = store_sum.sort_values('net_cost', ascending=False)
            
            short_name = store.replace('Thrive ', '').replace('Cannabis ', '')[:12]
            ws_store = wb.create_sheet(title=f"Rewards - {short_name}")
            
            ws_store.cell(row=1, column=1).value = f"{store}"
            ws_store.cell(row=1, column=1).font = SECTION_FONT
            
            total_net_store = store_sum['net_cost'].sum()
            ws_store.cell(row=2, column=1).value = f"Total Redemptions: {store_sum['redemptions'].sum()}  |  Net Cost: ${total_net_store:,.2f}"
            ws_store.cell(row=2, column=1).font = SUBTITLE_FONT
            
            cols = [
                ('reward_name', 'text', 'Reward Name'),
                ('redemptions', 'number', 'Redemptions'),
                ('units', 'number', 'Units'),
                ('retail_value', 'currency', 'Retail Value'),
                ('cost', 'currency', 'Product Cost'),
                ('net_cost', 'currency', 'Net Cost'),
            ]
            
            for col_num, (_, _, label) in enumerate(cols, 1):
                ws_store.cell(row=4, column=col_num).value = label
            format_header_row(ws_store, 4, len(cols))
            
            for row_num, (_, data) in enumerate(store_sum.iterrows(), 5):
                for col_num, (col_key, col_type, _) in enumerate(cols, 1):
                    hl = 'warning' if row_num == 5 else None
                    format_data_cell(ws_store, row_num, col_num, data.get(col_key, 0), col_type, highlight=hl)
            
            # Total
            row = len(store_sum) + 6
            format_data_cell(ws_store, row, 1, 'TOTAL', 'text', is_total=True)
            format_data_cell(ws_store, row, 2, store_sum['redemptions'].sum(), 'number', is_total=True)
            format_data_cell(ws_store, row, 3, store_sum['units'].sum(), 'number', is_total=True)
            format_data_cell(ws_store, row, 4, store_sum['retail_value'].sum(), 'currency', is_total=True)
            format_data_cell(ws_store, row, 5, store_sum['cost'].sum(), 'currency', is_total=True)
            format_data_cell(ws_store, row, 6, store_sum['net_cost'].sum(), 'currency', is_total=True)
            
            auto_column_width(ws_store)
    
    # Markouts by Employee
    if len(markouts_df) > 0:
        ws_mark = wb.create_sheet(title="Markouts by Employee")
        
        emp_stores = markouts_df.groupby(['customer_name', 'store_clean']).size().reset_index(name='count')
        primary_store = emp_stores.loc[emp_stores.groupby('customer_name')['count'].idxmax()][['customer_name', 'store_clean']]
        
        emp_products = markouts_df.groupby('customer_name')['product'].apply(
            lambda x: ', '.join(x.unique()[:3]) + ('...' if len(x.unique()) > 3 else '')
        ).reset_index()
        emp_products.columns = ['customer_name', 'products']
        
        emp_sum = markouts_df.groupby('customer_name').agg(
            redemptions=('receipt_id', 'count'),
            units=('quantity', 'sum'),
            cost=('cost', 'sum'),
        ).reset_index()
        emp_sum = emp_sum.merge(primary_store, on='customer_name', how='left')
        emp_sum = emp_sum.merge(emp_products, on='customer_name', how='left')
        emp_sum = emp_sum.sort_values('cost', ascending=False)
        emp_sum['rank'] = range(1, len(emp_sum) + 1)
        
        cols = [
            ('rank', 'number', 'Rank'),
            ('customer_name', 'text', 'Employee Name'),
            ('store_clean', 'text', 'Store'),
            ('redemptions', 'number', 'Redemptions'),
            ('units', 'number', 'Units'),
            ('cost', 'currency', 'Product Cost'),
            ('products', 'text', 'Products'),
        ]
        
        for col_num, (_, _, label) in enumerate(cols, 1):
            ws_mark.cell(row=1, column=col_num).value = label
        format_header_row(ws_mark, 1, len(cols))
        
        for row_num, (_, data) in enumerate(emp_sum.iterrows(), 2):
            for col_num, (col_key, col_type, _) in enumerate(cols, 1):
                hl = 'orange' if row_num <= 6 else None
                format_data_cell(ws_mark, row_num, col_num, data.get(col_key, 0), col_type, highlight=hl)
        
        auto_column_width(ws_mark)
        ws_mark.freeze_panes = 'A2'
    
    wb.save(output_path)
    return {'rewards_net': rewards_net, 'markouts_net': markouts_net, 'total_net': total_net, 'monthly': monthly}


# =============================================================================
# FILE MANAGEMENT
# =============================================================================

def setup_folders(base_folder):
    folders = {
        'inbox': base_folder / 'inbox',
        'archive': base_folder / 'archive',
        'reports': base_folder / 'reports',
    }
    for folder in folders.values():
        folder.mkdir(parents=True, exist_ok=True)
    return folders


def find_files(inbox_folder, patterns):
    files = {}
    used_files = set()
    check_order = ['bt_performance', 'customers', 'sales']
    
    for file_type in check_order:
        keywords = patterns[file_type]
        for csv_file in inbox_folder.glob('*.csv'):
            if csv_file in used_files:
                continue
            filename_lower = csv_file.name.lower()
            if any(kw in filename_lower for kw in keywords):
                files[file_type] = csv_file
                used_files.add(csv_file)
                break
    return files


def archive_files(files, archive_folder):
    import shutil
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    archive_subfolder = archive_folder / timestamp
    archive_subfolder.mkdir(exist_ok=True)
    
    for file_type, filepath in files.items():
        if filepath and filepath.exists():
            shutil.move(str(filepath), str(archive_subfolder / filepath.name))
            print(f"  📁 Archived: {filepath.name}")


# =============================================================================
# CONFIGURATION
# =============================================================================

CONFIG = {
    'base_folder': Path.home() / 'Desktop' / 'Thrive Analytics',
    'file_patterns': {
        'sales': ['margin', 'line_item', 'john'],
        'bt_performance': ['bt sales', 'bt_sales', 'budtender'],
        'customers': ['customer'],
    }
}


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("\n" + "="*70)
    print("  THRIVE CANNABIS - MASTER ANALYTICS SUITE v2")
    print("="*70)
    print(f"  Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    base_folder = CONFIG['base_folder']
    folders = setup_folders(base_folder)
    
    print(f"\n📁 Watching folder: {folders['inbox']}")
    
    files = find_files(folders['inbox'], CONFIG['file_patterns'])
    
    if 'sales' not in files:
        print("\n⚠️  No sales file found in inbox!")
        print(f"   Please drop your Flowhub 'Margin Report' CSV in:")
        print(f"   {folders['inbox']}")
        return
    
    print(f"\n📄 Found files:")
    for file_type, filepath in files.items():
        print(f"   {file_type}: {filepath.name}")
    
    # Load data
    print(f"\n📊 Loading data...")
    sales_df = load_sales_data(files['sales'])
    bt_df = load_bt_performance(files['bt_performance']) if 'bt_performance' in files else None
    cust_attr_df = load_customer_attributes(files['customers']) if 'customers' in files else None
    
    print(f"   ✓ {len(sales_df):,} line items loaded")
    
    # Get date range
    valid_dates = sales_df['sale_date'].dropna()
    date_range = f"{valid_dates.min()} to {valid_dates.max()}" if len(valid_dates) > 0 else 'N/A'
    
    # Classify transactions
    sales_df['transaction_type'] = sales_df.apply(classify_transaction, axis=1)
    regular_df = sales_df[sales_df['transaction_type'] == 'REGULAR'].copy()
    regular_df['deal_type'] = regular_df.apply(classify_deal_type, axis=1)
    
    print(f"   ✓ {len(regular_df):,} regular sales")
    print(f"   ✓ Date range: {date_range}")
    
    # Create output folder
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_folder = folders['reports'] / timestamp
    output_folder.mkdir(exist_ok=True)
    
    # Generate reports
    print(f"\n📊 Generating Reports...")
    
    margin_totals = create_margin_report(regular_df, output_folder / "Margin_Report.xlsx", date_range)
    print(f"   ✓ Margin_Report.xlsx")
    
    create_deal_report(regular_df, output_folder / "Deal_Performance_Report.xlsx", date_range)
    print(f"   ✓ Deal_Performance_Report.xlsx (with legend + Top 10 by store)")
    
    if bt_df is not None:
        create_budtender_report(bt_df, sales_df, output_folder / "Budtender_Performance_Report.xlsx", date_range)
        print(f"   ✓ Budtender_Performance_Report.xlsx")
    
    create_customer_report(sales_df, cust_attr_df, output_folder / "Customer_Insights_Report.xlsx", date_range)
    print(f"   ✓ Customer_Insights_Report.xlsx (PERIOD-SPECIFIC metrics)")
    
    rewards_totals = create_rewards_report(sales_df, output_folder / "Rewards_Markout_Report.xlsx", date_range)
    print(f"   ✓ Rewards_Markout_Report.xlsx (actual reward names by store)")
    
    # Summary
    print(f"\n" + "="*70)
    print("  EXECUTIVE SUMMARY")
    print("="*70)
    print(f"\n  📅 Period: {date_range}")
    print(f"\n  💰 MARGIN PERFORMANCE")
    print(f"     Total Revenue:     ${margin_totals['total_revenue']:,.2f}")
    print(f"     Full Price Sales:  ${margin_totals['full_price_sales']:,.2f} ({margin_totals['pct_full_price']}%)")
    print(f"     Blended Margin:    {margin_totals['blended_margin']}%")
    
    print(f"\n  🎁 REWARDS & MARKOUTS")
    print(f"     Total Net Cost:    ${rewards_totals['total_net']:,.2f}")
    print(f"     Monthly Projection: ${rewards_totals['monthly']:,.2f}")
    
    # Archive
    print(f"\n📁 Archiving processed files...")
    archive_files(files, folders['archive'])
    
    print("\n" + "="*70)
    print(f"  ✅ COMPLETE!")
    print(f"  📂 Reports saved to: {output_folder}")
    print("="*70 + "\n")


if __name__ == "__main__":
    main()
