"""
Single source of truth for all Excel colors, fonts, fills, borders, alignments.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# Color constants
# ---------------------------------------------------------------------------
THRIVE_GREEN = "2E7D32"
DARK_GREEN = "1B5E20"
LIGHT_GREEN = "E8F5E9"
HEADER_BG = "1B5E20"
ALTERNATE_ROW = "F5F5F5"
WHITE = "FFFFFF"
BLACK = "000000"
GOLD = "FFD700"
LIGHT_GOLD = "FFF8DC"
RED = "D32F2F"
LIGHT_RED = "FFEBEE"
ORANGE = "FF9800"
LIGHT_ORANGE = "FFF3E0"
TOTAL_ROW_BG = "E3F2FD"
GRAY_666 = "666666"

# ---------------------------------------------------------------------------
# Fonts
# ---------------------------------------------------------------------------
TITLE_FONT = Font(name="Calibri", size=24, bold=True, color=DARK_GREEN)
SUBTITLE_FONT = Font(name="Calibri", size=12, italic=True, color=GRAY_666)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
DATA_FONT = Font(name="Calibri", size=10, color=BLACK)
TOTAL_FONT = Font(name="Calibri", size=10, bold=True, color=BLACK)
SECTION_FONT = Font(name="Calibri", size=14, bold=True, color=DARK_GREEN)
KPI_VALUE_FONT = Font(name="Calibri", size=28, bold=True, color=DARK_GREEN)
KPI_LABEL_FONT = Font(name="Calibri", size=10, color=GRAY_666)
POSITIVE_KPI_FONT = Font(name="Calibri", size=28, bold=True, color=THRIVE_GREEN)
NEGATIVE_KPI_FONT = Font(name="Calibri", size=28, bold=True, color=RED)
INSIGHT_TITLE_FONT = Font(name="Calibri", size=11, bold=True)
INSIGHT_BODY_FONT = Font(name="Calibri", size=10, italic=True)
REC_TITLE_FONT = Font(name="Calibri", size=12, bold=True)
REC_BODY_FONT = Font(name="Calibri", size=10)
LEGEND_BOLD_FONT = Font(name="Calibri", size=10, bold=True)

# ---------------------------------------------------------------------------
# Fills
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
ALTERNATE_FILL = PatternFill(start_color=ALTERNATE_ROW, end_color=ALTERNATE_ROW, fill_type="solid")
TOTAL_FILL = PatternFill(start_color=TOTAL_ROW_BG, end_color=TOTAL_ROW_BG, fill_type="solid")
GOLD_FILL = PatternFill(start_color=LIGHT_GOLD, end_color=LIGHT_GOLD, fill_type="solid")
WARNING_FILL = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
ORANGE_FILL = PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type="solid")

# ---------------------------------------------------------------------------
# Borders
# ---------------------------------------------------------------------------
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
HEADER_BORDER = Border(
    left=Side(style="thin", color=DARK_GREEN),
    right=Side(style="thin", color=DARK_GREEN),
    top=Side(style="thin", color=DARK_GREEN),
    bottom=Side(style="medium", color=DARK_GREEN),
)
TOTAL_BORDER = Border(
    left=Side(style="thin", color="999999"),
    right=Side(style="thin", color="999999"),
    top=Side(style="medium", color="999999"),
    bottom=Side(style="medium", color="999999"),
)

# ---------------------------------------------------------------------------
# Alignments
# ---------------------------------------------------------------------------
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ---------------------------------------------------------------------------
# Highlight name → fill mapping
# ---------------------------------------------------------------------------
HIGHLIGHT_FILLS = {
    "gold": GOLD_FILL,
    "warning": WARNING_FILL,
    "orange": ORANGE_FILL,
    "green": LIGHT_GREEN_FILL,
}
