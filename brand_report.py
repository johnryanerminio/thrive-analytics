#!/usr/bin/env python3
"""
================================================================================
THRIVE CANNABIS - BRAND PERFORMANCE REPORT GENERATOR v2
================================================================================
Generate detailed performance reports for specific brands to use in 
vendor negotiations, pricing discussions, and performance reviews.

NEW IN V2:
  - Product Type breakdown (Flower, Vapes, Pre-Rolls, etc.)
  - Category margin comparison (vs category avg, not company avg)
  - Brand ranking within each category
  - Top Products by store tabs

USAGE:
  python3 brand_report.py "WYLD"                    # Single brand
  python3 brand_report.py "WYLD" "STIIIZY" "FADE"   # Multiple brands
  python3 brand_report.py --list                    # List available brands
  python3 brand_report.py --top 10                  # Top 10 by revenue

OUTPUT:
  Reports saved to: ~/Desktop/Thrive Analytics/brand_reports/

================================================================================
"""

import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import argparse
import warnings
warnings.filterwarnings('ignore')


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

DARK_GREEN = "1B5E20"
LIGHT_GREEN = "E8F5E9"
HEADER_BG = "1B5E20"
ALTERNATE_ROW = "F5F5F5"
WHITE = "FFFFFF"
BLACK = "000000"
TOTAL_ROW_BG = "E3F2FD"
LIGHT_GOLD = "FFF8DC"
LIGHT_RED = "FFEBEE"

TITLE_FONT = Font(name='Calibri', size=24, bold=True, color=DARK_GREEN)
SUBTITLE_FONT = Font(name='Calibri', size=12, italic=True, color="666666")
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color=WHITE)
DATA_FONT = Font(name='Calibri', size=10, color=BLACK)
TOTAL_FONT = Font(name='Calibri', size=10, bold=True, color=BLACK)
SECTION_FONT = Font(name='Calibri', size=14, bold=True, color=DARK_GREEN)
KPI_VALUE_FONT = Font(name='Calibri', size=28, bold=True, color=DARK_GREEN)
KPI_LABEL_FONT = Font(name='Calibri', size=10, color="666666")

HEADER_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
ALTERNATE_FILL = PatternFill(start_color=ALTERNATE_ROW, end_color=ALTERNATE_ROW, fill_type="solid")
TOTAL_FILL = PatternFill(start_color=TOTAL_ROW_BG, end_color=TOTAL_ROW_BG, fill_type="solid")
GOLD_FILL = PatternFill(start_color=LIGHT_GOLD, end_color=LIGHT_GOLD, fill_type="solid")
WARNING_FILL = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
HEADER_BORDER = Border(
    left=Side(style='thin', color=DARK_GREEN),
    right=Side(style='thin', color=DARK_GREEN),
    top=Side(style='thin', color=DARK_GREEN),
    bottom=Side(style='medium', color=DARK_GREEN)
)
TOTAL_BORDER = Border(
    left=Side(style='thin', color='999999'),
    right=Side(style='thin', color='999999'),
    top=Side(style='medium', color='999999'),
    bottom=Side(style='medium', color='999999')
)

CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')


def format_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = HEADER_BORDER


def format_data_cell(ws, row_num, col_num, value, col_type='text', is_total=False, highlight=None):
    cell = ws.cell(row=row_num, column=col_num)
    cell.value = value
    cell.font = TOTAL_FONT if is_total else DATA_FONT
    cell.border = TOTAL_BORDER if is_total else THIN_BORDER
    cell.alignment = RIGHT if col_type in ['currency', 'number', 'percent', 'decimal'] else LEFT
    
    if col_type == 'currency':
        cell.number_format = '"$"#,##0.00'
    elif col_type == 'percent':
        cell.number_format = '0.0"%"'
    elif col_type == 'number':
        cell.number_format = '#,##0'
    
    if highlight == 'gold':
        cell.fill = GOLD_FILL
    elif highlight == 'warning':
        cell.fill = WARNING_FILL
    elif is_total:
        cell.fill = TOTAL_FILL
    elif row_num % 2 == 0:
        cell.fill = ALTERNATE_FILL


def auto_column_width(ws, min_width=10, max_width=50):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def add_kpi_card(ws, row, col, value, label, format_type='currency'):
    value_cell = ws.cell(row=row, column=col)
    label_cell = ws.cell(row=row + 1, column=col)
    
    value_cell.value = value
    value_cell.font = KPI_VALUE_FONT
    value_cell.alignment = CENTER
    
    if format_type == 'currency':
        value_cell.number_format = '"$"#,##0'
    elif format_type == 'percent':
        value_cell.number_format = '0.0"%"'
    elif format_type == 'number':
        value_cell.number_format = '#,##0'
    elif format_type == 'text':
        value_cell.number_format = '@'
    
    label_cell.value = label
    label_cell.font = KPI_LABEL_FONT
    label_cell.alignment = CENTER


def find_sales_file(base_folder):
    """Find the most recent sales CSV, searching inbox (including year subdirs) then archive."""
    inbox = base_folder / 'inbox'
    archive = base_folder / 'archive'
    keywords = ['margin', 'john', 'line_item']

    # Search inbox recursively (handles inbox/*.csv AND inbox/2025/*.csv etc.)
    if inbox.exists():
        matches = []
        for csv_file in inbox.rglob('*.csv'):
            filename_lower = csv_file.name.lower()
            if any(kw in filename_lower for kw in keywords):
                matches.append(csv_file)
        if matches:
            # Return most recently modified file
            return max(matches, key=lambda f: f.stat().st_mtime)

    if archive.exists():
        archive_folders = sorted(archive.iterdir(), reverse=True)
        for folder in archive_folders:
            if folder.is_dir():
                for csv_file in folder.glob('*.csv'):
                    filename_lower = csv_file.name.lower()
                    if any(kw in filename_lower for kw in keywords):
                        return csv_file
    return None


def load_sales_data(filepath):
    df = pd.read_csv(filepath)
    
    col_map = {
        'Pre-Discount, Pre-Tax Total': 'pre_discount_revenue',
        'Discounts': 'discounts',
        'Post-Discount, Pre-Tax Total': 'actual_revenue',
        'Net Profit': 'net_profit',
        'Cost': 'cost',
        'Quantity Sold': 'quantity',
        'Receipt ID': 'receipt_id',
        'Completed At': 'completed_at',
        'Store': 'store',
        'Product': 'product',
        'Brand': 'brand',
        'Variant Type': 'category',
        'Deals Used': 'deals_used',
    }
    df = df.rename(columns=col_map)
    
    currency_cols = ['pre_discount_revenue', 'discounts', 'actual_revenue', 'net_profit', 'cost']
    for col in currency_cols:
        if col in df.columns:
            df[col] = df[col].replace(r'[\$,]', '', regex=True).astype(float)
    
    df['completed_at'] = pd.to_datetime(df['completed_at'], format='%m/%d/%Y %I:%M:%S %p', errors='coerce')
    df['sale_date'] = df['completed_at'].dt.date
    df = df.dropna(subset=['completed_at'])
    
    df['store_clean'] = df['store'].str.replace(r' - RD\d+', '', regex=True).str.strip()
    df['brand_clean'] = df['brand'].str.strip()
    df['category_clean'] = df['category'].str.strip().str.upper() if 'category' in df.columns else 'UNKNOWN'

    # Normalize category names (handle casing variants and common aliases)
    category_normalization = {
        'ACCESSORY': 'ACCESSORY',
        'PRE-ROLL': 'PRE ROLL',
        'PREROLL': 'PRE ROLL',
        'PRE-ROLLS': 'PRE ROLL',
        'PRE ROLLS': 'PRE ROLL',
        'PRE-ROLL PACK': 'PRE ROLL PACK',
        'PREROLL PACK': 'PRE ROLL PACK',
        'VAPE': 'CARTRIDGE',
        'CART': 'CARTRIDGE',
        'CARTS': 'CARTRIDGE',
        'DISPOSABLE': 'DISPOSABLE VAPE',
        'DISPO': 'DISPOSABLE VAPE',
        'GUMMY': 'EDIBLE',
        'GUMMIES': 'EDIBLE',
        'EDIBLES': 'EDIBLE',
    }
    df['category_clean'] = df['category_clean'].replace(category_normalization)
    df['deals_upper'] = df['deals_used'].fillna('').str.upper()
    df['has_discount'] = df['discounts'] > 0
    
    def is_regular(row):
        deals = str(row.get('deals_upper', ''))
        rev = row.get('actual_revenue', 0)
        if 'REWARD' in deals or 'POINT' in deals or 'MARKOUT' in deals or 'MARK OUT' in deals:
            return False
        if rev <= 1.00:
            return False
        return True
    
    df['is_regular'] = df.apply(is_regular, axis=1)
    return df


def create_brand_report(brand_df, brand_name, output_path, date_range, 
                        category_margin_lookup, brand_category_rankings):
    wb = Workbook()
    
    # Calculate totals
    total_units = brand_df['quantity'].sum()
    total_revenue = brand_df['actual_revenue'].sum()
    total_cost = brand_df['cost'].sum()
    total_discounts = brand_df['discounts'].sum()
    total_profit = brand_df['net_profit'].sum()
    
    full_price_df = brand_df[~brand_df['has_discount']]
    discounted_df = brand_df[brand_df['has_discount']]
    
    fp_revenue = full_price_df['actual_revenue'].sum()
    fp_cost = full_price_df['cost'].sum()
    disc_revenue = discounted_df['actual_revenue'].sum()
    disc_cost = discounted_df['cost'].sum()
    
    overall_margin = (total_revenue - total_cost) / total_revenue * 100 if total_revenue > 0 else 0
    fp_margin = (fp_revenue - fp_cost) / fp_revenue * 100 if fp_revenue > 0 else 0
    disc_margin = (disc_revenue - disc_cost) / disc_revenue * 100 if disc_revenue > 0 else 0
    pct_full_price = fp_revenue / total_revenue * 100 if total_revenue > 0 else 0
    avg_discount_rate = total_discounts / (total_revenue + total_discounts) * 100 if (total_revenue + total_discounts) > 0 else 0
    
    # By product type
    brand_categories = brand_df.groupby('category_clean').agg(
        units=('quantity', 'sum'),
        revenue=('actual_revenue', 'sum'),
        cost=('cost', 'sum'),
        profit=('net_profit', 'sum'),
        discounts=('discounts', 'sum'),
    ).reset_index()
    
    brand_categories['margin'] = ((brand_categories['revenue'] - brand_categories['cost']) / 
                                   brand_categories['revenue'] * 100).round(1)
    brand_categories['category_avg_margin'] = brand_categories['category_clean'].map(category_margin_lookup)
    brand_categories['vs_category'] = (brand_categories['margin'] - brand_categories['category_avg_margin']).round(1)
    
    for idx, row in brand_categories.iterrows():
        cat = row['category_clean']
        rank_data = brand_category_rankings[(brand_category_rankings['category_clean'] == cat) & 
                                             (brand_category_rankings['brand_clean'] == brand_name)]
        if len(rank_data) > 0:
            brand_categories.loc[idx, 'rank'] = int(rank_data['rank'].values[0])
            brand_categories.loc[idx, 'total_brands'] = int(rank_data['total_brands'].values[0])
        else:
            brand_categories.loc[idx, 'rank'] = 0
            brand_categories.loc[idx, 'total_brands'] = 0
    
    brand_categories = brand_categories.sort_values('revenue', ascending=False)
    
    if len(brand_categories) > 0:
        primary_category = brand_categories.iloc[0]['category_clean']
        primary_cat_margin = category_margin_lookup.get(primary_category, overall_margin)
        primary_rank = int(brand_categories.iloc[0].get('rank', 0))
        primary_total = int(brand_categories.iloc[0].get('total_brands', 0))
    else:
        primary_category = 'UNKNOWN'
        primary_cat_margin = overall_margin
        primary_rank = 0
        primary_total = 0
    
    margin_vs_cat = overall_margin - primary_cat_margin
    
    # === Executive Summary ===
    ws = wb.active
    ws.title = "Executive Summary"
    
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    ws.cell(row=1, column=1).value = f"{brand_name.upper()}"
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    
    ws.cell(row=2, column=1).value = f"Brand Performance Report  |  {date_range}  |  Prepared by Thrive Cannabis"
    ws.cell(row=2, column=1).font = SUBTITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    
    ws.cell(row=4, column=1).value = "SALES PERFORMANCE"
    ws.cell(row=4, column=1).font = SECTION_FONT
    
    add_kpi_card(ws, 6, 1, total_units, "UNITS SOLD", 'number')
    add_kpi_card(ws, 6, 3, total_revenue, "TOTAL REVENUE", 'currency')
    add_kpi_card(ws, 6, 5, total_profit, "NET PROFIT", 'currency')
    add_kpi_card(ws, 6, 7, overall_margin, "OVERALL MARGIN", 'percent')
    
    ws.cell(row=10, column=1).value = "PRICING ANALYSIS"
    ws.cell(row=10, column=1).font = SECTION_FONT
    
    add_kpi_card(ws, 12, 1, pct_full_price, "% SOLD FULL PRICE", 'percent')
    add_kpi_card(ws, 12, 3, fp_margin, "FULL PRICE MARGIN", 'percent')
    add_kpi_card(ws, 12, 5, disc_margin, "DISCOUNTED MARGIN", 'percent')
    add_kpi_card(ws, 12, 7, avg_discount_rate, "AVG DISCOUNT RATE", 'percent')
    
    ws.cell(row=16, column=1).value = "CATEGORY RANKING"
    ws.cell(row=16, column=1).font = SECTION_FONT
    
    rank_text = f"#{primary_rank} of {primary_total}" if primary_rank > 0 else "N/A"
    add_kpi_card(ws, 18, 1, rank_text, f"RANK IN {primary_category}", 'text')
    add_kpi_card(ws, 18, 3, primary_cat_margin, f"{primary_category} AVG MARGIN", 'percent')
    
    ws.cell(row=18, column=5).value = margin_vs_cat
    ws.cell(row=18, column=5).font = KPI_VALUE_FONT
    ws.cell(row=18, column=5).number_format = '+0.0;-0.0;0.0'
    if margin_vs_cat >= 0:
        ws.cell(row=18, column=5).font = Font(name='Calibri', size=28, bold=True, color="2E7D32")
    else:
        ws.cell(row=18, column=5).font = Font(name='Calibri', size=28, bold=True, color="D32F2F")
    ws.cell(row=19, column=5).value = "PTS VS CATEGORY"
    ws.cell(row=19, column=5).font = KPI_LABEL_FONT
    ws.cell(row=19, column=5).alignment = CENTER
    
    ws.cell(row=22, column=1).value = "💡 KEY INSIGHT:"
    ws.cell(row=22, column=1).font = Font(name='Calibri', size=11, bold=True)
    
    if margin_vs_cat < -10:
        insight = f"{brand_name} margin ({overall_margin:.1f}%) is {abs(margin_vs_cat):.0f} pts BELOW {primary_category} average ({primary_cat_margin:.1f}%). Strong case for cost reduction."
    elif margin_vs_cat < 0:
        insight = f"{brand_name} margin ({overall_margin:.1f}%) is slightly below {primary_category} average ({primary_cat_margin:.1f}%). Room for negotiation."
    elif pct_full_price < 30:
        insight = f"Only {pct_full_price:.0f}% sells at full price. Heavy discounting required to move inventory."
    else:
        insight = f"{brand_name} performs well - ranked #{primary_rank} in {primary_category} with {overall_margin:.1f}% margin vs {primary_cat_margin:.1f}% category average."
    
    ws.cell(row=23, column=1).value = insight
    ws.cell(row=23, column=1).font = Font(name='Calibri', size=10, italic=True)
    ws.merge_cells(start_row=23, start_column=1, end_row=23, end_column=8)
    
    # === By Product Type ===
    if len(brand_categories) > 1:
        ws2 = wb.create_sheet(title="By Product Type")
        ws2.cell(row=1, column=1).value = f"{brand_name} Performance by Product Type"
        ws2.cell(row=1, column=1).font = SECTION_FONT
        
        cols = [
            ('category_clean', 'text', 'Product Type'),
            ('rank', 'number', 'Category Rank'),
            ('total_brands', 'number', 'Brands in Category'),
            ('units', 'number', 'Units'),
            ('revenue', 'currency', 'Revenue'),
            ('margin', 'percent', 'Margin'),
            ('category_avg_margin', 'percent', 'Category Avg'),
            ('vs_category', 'percent', 'vs Category'),
            ('profit', 'currency', 'Net Profit'),
        ]
        
        for col_num, (_, _, label) in enumerate(cols, 1):
            ws2.cell(row=3, column=col_num).value = label
        format_header_row(ws2, 3, len(cols))
        
        for row_num, (_, data) in enumerate(brand_categories.iterrows(), 4):
            for col_num, (col_key, col_type, _) in enumerate(cols, 1):
                val = data.get(col_key, 0)
                vs_cat = data.get('vs_category', 0)
                highlight = 'gold' if col_key == 'vs_category' and vs_cat >= 0 else ('warning' if col_key == 'vs_category' and vs_cat < 0 else None)
                format_data_cell(ws2, row_num, col_num, val if not pd.isna(val) else 0, col_type, highlight=highlight)
        
        auto_column_width(ws2)
        ws2.freeze_panes = 'A4'
    
    # === Deal Performance ===
    ws3 = wb.create_sheet(title="Deal Performance")
    
    def extract_deals(deals_str):
        if pd.isna(deals_str) or deals_str == '':
            return []
        return [d.strip() for d in str(deals_str).split(',') if d.strip()]
    
    deals_expanded = []
    for _, row in brand_df.iterrows():
        deals = extract_deals(row['deals_used'])
        if deals:
            n_deals = len(deals)
            for deal in deals:
                deals_expanded.append({
                    'deal_name': deal, 'receipt_id': row['receipt_id'],
                    'revenue': row['actual_revenue'] / n_deals,
                    'discounts': row['discounts'] / n_deals,
                    'quantity': row['quantity'] / n_deals,
                    'cost': row['cost'] / n_deals,
                    'profit': row['net_profit'] / n_deals,
                })
    
    if deals_expanded:
        deals_df = pd.DataFrame(deals_expanded)
        deal_summary = deals_df.groupby('deal_name').agg(
            times_used=('receipt_id', 'nunique'), units=('quantity', 'sum'),
            revenue=('revenue', 'sum'), discounts=('discounts', 'sum'),
            cost=('cost', 'sum'), profit=('profit', 'sum'),
        ).reset_index()
        deal_summary['margin'] = ((deal_summary['revenue'] - deal_summary['cost']) / deal_summary['revenue'] * 100).round(1)
        deal_summary['avg_discount'] = (deal_summary['discounts'] / (deal_summary['revenue'] + deal_summary['discounts']) * 100).round(1)
        deal_summary = deal_summary.sort_values('times_used', ascending=False)
        
        ws3.cell(row=1, column=1).value = f"Deals Used with {brand_name} Products"
        ws3.cell(row=1, column=1).font = SECTION_FONT
        
        cols = [('deal_name', 'text', 'Deal Name'), ('times_used', 'number', 'Times Used'),
                ('units', 'number', 'Units'), ('revenue', 'currency', 'Revenue'),
                ('discounts', 'currency', 'Discounts'), ('avg_discount', 'percent', 'Discount %'),
                ('margin', 'percent', 'Margin'), ('profit', 'currency', 'Net Profit')]
        
        for col_num, (_, _, label) in enumerate(cols, 1):
            ws3.cell(row=3, column=col_num).value = label
        format_header_row(ws3, 3, len(cols))
        
        for row_num, (_, data) in enumerate(deal_summary.iterrows(), 4):
            for col_num, (col_key, col_type, _) in enumerate(cols, 1):
                val = data.get(col_key, 0)
                highlight = 'warning' if data.get('margin', 100) < 40 else ('gold' if row_num <= 6 else None)
                format_data_cell(ws3, row_num, col_num, val if not pd.isna(val) else 0, col_type, highlight=highlight)
        
        auto_column_width(ws3)
        ws3.freeze_panes = 'A4'
    
    # === Top Products ===
    ws4 = wb.create_sheet(title="Top Products")
    product_summary = brand_df.groupby('product').agg(
        units=('quantity', 'sum'), revenue=('actual_revenue', 'sum'),
        cost=('cost', 'sum'), profit=('net_profit', 'sum'), transactions=('receipt_id', 'nunique'),
    ).reset_index()
    product_summary['margin'] = ((product_summary['revenue'] - product_summary['cost']) / product_summary['revenue'] * 100).round(1)
    product_summary = product_summary.sort_values('revenue', ascending=False).head(25)
    
    ws4.cell(row=1, column=1).value = f"Top 25 {brand_name} Products (All Stores)"
    ws4.cell(row=1, column=1).font = SECTION_FONT
    
    prod_cols = [('product', 'text', 'Product'), ('transactions', 'number', 'Transactions'),
                 ('units', 'number', 'Units'), ('revenue', 'currency', 'Revenue'),
                 ('margin', 'percent', 'Margin'), ('profit', 'currency', 'Net Profit')]
    
    for col_num, (_, _, label) in enumerate(prod_cols, 1):
        ws4.cell(row=3, column=col_num).value = label
    format_header_row(ws4, 3, len(prod_cols))
    
    for row_num, (_, data) in enumerate(product_summary.iterrows(), 4):
        for col_num, (col_key, col_type, _) in enumerate(prod_cols, 1):
            highlight = 'gold' if row_num <= 8 else None
            format_data_cell(ws4, row_num, col_num, data.get(col_key, 0), col_type, highlight=highlight)
    
    auto_column_width(ws4)
    ws4.freeze_panes = 'A4'
    
    # === Products by Store ===
    stores = sorted(brand_df['store_clean'].dropna().unique())
    for store in stores:
        store_df = brand_df[brand_df['store_clean'] == store]
        store_products = store_df.groupby('product').agg(
            units=('quantity', 'sum'), revenue=('actual_revenue', 'sum'),
            cost=('cost', 'sum'), profit=('net_profit', 'sum'), transactions=('receipt_id', 'nunique'),
        ).reset_index()
        store_products['margin'] = ((store_products['revenue'] - store_products['cost']) / store_products['revenue'] * 100).round(1)
        store_products = store_products.sort_values('revenue', ascending=False).head(15)
        
        if len(store_products) == 0:
            continue
        
        short_name = store.replace('Thrive ', '').replace('Cannabis ', '')[:12]
        ws_store = wb.create_sheet(title=f"Products - {short_name}")
        ws_store.cell(row=1, column=1).value = f"Top {brand_name} Products at {store}"
        ws_store.cell(row=1, column=1).font = SECTION_FONT
        
        for col_num, (_, _, label) in enumerate(prod_cols, 1):
            ws_store.cell(row=3, column=col_num).value = label
        format_header_row(ws_store, 3, len(prod_cols))
        
        for row_num, (_, data) in enumerate(store_products.iterrows(), 4):
            for col_num, (col_key, col_type, _) in enumerate(prod_cols, 1):
                highlight = 'gold' if row_num <= 6 else None
                format_data_cell(ws_store, row_num, col_num, data.get(col_key, 0), col_type, highlight=highlight)
        
        auto_column_width(ws_store)
        ws_store.freeze_panes = 'A4'
    
    # === By Store ===
    ws_stores = wb.create_sheet(title="By Store")
    store_summary = brand_df.groupby('store_clean').agg(
        units=('quantity', 'sum'), revenue=('actual_revenue', 'sum'),
        cost=('cost', 'sum'), profit=('net_profit', 'sum'), discounts=('discounts', 'sum'),
    ).reset_index()
    store_summary['margin'] = ((store_summary['revenue'] - store_summary['cost']) / store_summary['revenue'] * 100).round(1)
    store_summary['discount_rate'] = (store_summary['discounts'] / (store_summary['revenue'] + store_summary['discounts']) * 100).round(1)
    store_summary = store_summary.sort_values('revenue', ascending=False)
    
    ws_stores.cell(row=1, column=1).value = f"{brand_name} Performance by Store"
    ws_stores.cell(row=1, column=1).font = SECTION_FONT
    
    store_cols = [('store_clean', 'text', 'Store'), ('units', 'number', 'Units'),
                  ('revenue', 'currency', 'Revenue'), ('margin', 'percent', 'Margin'),
                  ('discounts', 'currency', 'Discounts'), ('discount_rate', 'percent', 'Discount Rate'),
                  ('profit', 'currency', 'Net Profit')]
    
    for col_num, (_, _, label) in enumerate(store_cols, 1):
        ws_stores.cell(row=3, column=col_num).value = label
    format_header_row(ws_stores, 3, len(store_cols))
    
    for row_num, (_, data) in enumerate(store_summary.iterrows(), 4):
        for col_num, (col_key, col_type, _) in enumerate(store_cols, 1):
            format_data_cell(ws_stores, row_num, col_num, data.get(col_key, 0), col_type)
    
    auto_column_width(ws_stores)
    ws_stores.freeze_panes = 'A4'
    
    # === Recommendations ===
    ws_rec = wb.create_sheet(title="Recommendations")
    ws_rec.cell(row=1, column=1).value = "PRICING & PROMOTION RECOMMENDATIONS"
    ws_rec.cell(row=1, column=1).font = TITLE_FONT
    ws_rec.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    
    recommendations = []
    
    for _, cat_row in brand_categories.iterrows():
        cat = cat_row['category_clean']
        vs_cat = cat_row.get('vs_category', 0)
        cat_margin = cat_row.get('margin', 0)
        cat_rank = cat_row.get('rank', 0)
        cat_total = cat_row.get('total_brands', 0)
        if pd.notna(vs_cat) and vs_cat < -10:
            recommendations.append((f"🔴 {cat}: BELOW CATEGORY AVERAGE",
                f"Margin ({cat_margin:.1f}%) is {abs(vs_cat):.0f} pts below {cat} average. Ranked #{int(cat_rank)} of {int(cat_total)}. Strong case for cost negotiation."))
    
    if margin_vs_cat < -5:
        recommendations.append(("🟡 CATEGORY BENCHMARK GAP",
            f"Overall margin ({overall_margin:.1f}%) trails {primary_category} average ({primary_cat_margin:.1f}%) by {abs(margin_vs_cat):.1f} pts."))
    
    if pct_full_price < 25:
        recommendations.append(("🔴 HIGH PROMOTION DEPENDENCY", f"Only {pct_full_price:.0f}% sells at full price."))
    
    if disc_margin < 35 and disc_revenue > 0:
        recommendations.append(("🔴 LOW DISCOUNTED MARGIN", f"Margin drops to {disc_margin:.1f}% when discounted."))
    
    if margin_vs_cat >= 5 and primary_rank <= 5 and primary_rank > 0:
        recommendations.append(("🟢 STRONG PERFORMER", f"Ranked #{primary_rank} in {primary_category} with {margin_vs_cat:.1f} pts above average."))
    
    if total_revenue > 5000:
        recommendations.append(("💰 VOLUME LEVERAGE", f"${total_revenue:,.0f} revenue provides negotiating leverage."))
    
    row = 5
    for title, detail in recommendations:
        ws_rec.cell(row=row, column=1).value = title
        ws_rec.cell(row=row, column=1).font = Font(name='Calibri', size=12, bold=True)
        ws_rec.cell(row=row+1, column=1).value = detail
        ws_rec.cell(row=row+1, column=1).font = Font(name='Calibri', size=10)
        ws_rec.cell(row=row+1, column=1).alignment = Alignment(wrap_text=True)
        ws_rec.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=6)
        row += 3
    
    ws_rec.column_dimensions['A'].width = 80
    
    wb.save(output_path)
    
    return {
        'revenue': total_revenue, 'margin': overall_margin, 'pct_full_price': pct_full_price,
        'units': total_units, 'primary_category': primary_category,
        'category_rank': primary_rank, 'category_total': primary_total, 'vs_category': margin_vs_cat,
    }


def main():
    parser = argparse.ArgumentParser(description='Generate brand performance reports')
    parser.add_argument('brands', nargs='*', help='Brand name(s)')
    parser.add_argument('--list', action='store_true', help='List brands')
    parser.add_argument('--top', type=int, metavar='N', help='Top N brands')
    args = parser.parse_args()
    
    base_folder = Path.home() / 'Desktop' / 'Thrive Analytics'
    output_folder = base_folder / 'brand_reports'
    output_folder.mkdir(parents=True, exist_ok=True)
    
    print("\n" + "="*70)
    print("  THRIVE CANNABIS - BRAND PERFORMANCE REPORT GENERATOR v2")
    print("="*70)
    
    sales_file = find_sales_file(base_folder)
    if not sales_file:
        print(f"\n⚠️  No sales data found in: {base_folder / 'inbox'}")
        return
    
    print(f"\n📄 Using: {sales_file.name}")
    df = load_sales_data(sales_file)
    regular_df = df[df['is_regular']].copy()
    
    valid_dates = df['sale_date'].dropna()
    date_range = f"{valid_dates.min()} to {valid_dates.max()}"
    
    # Category metrics
    cat_metrics = regular_df.groupby('category_clean').agg(
        total_revenue=('actual_revenue', 'sum'), total_cost=('cost', 'sum')).reset_index()
    cat_metrics['category_margin'] = ((cat_metrics['total_revenue'] - cat_metrics['total_cost']) / cat_metrics['total_revenue'] * 100).round(1)
    category_margin_lookup = dict(zip(cat_metrics['category_clean'], cat_metrics['category_margin']))
    
    # Brand rankings
    brand_cat_rev = regular_df.groupby(['category_clean', 'brand_clean'])['actual_revenue'].sum().reset_index()
    brand_cat_rev['rank'] = brand_cat_rev.groupby('category_clean')['actual_revenue'].rank(ascending=False, method='min')
    brand_cat_rev['total_brands'] = brand_cat_rev.groupby('category_clean')['brand_clean'].transform('count')
    
    brand_revenue = regular_df.groupby('brand_clean')['actual_revenue'].sum().sort_values(ascending=False)
    all_brands = brand_revenue.index.tolist()
    
    if args.list:
        print(f"\n📋 BRANDS ({len(all_brands)}):\n")
        for i, b in enumerate(all_brands[:50], 1):
            print(f"{i:<4}{b[:40]:<42}${brand_revenue[b]:>12,.2f}")
        return
    
    brands_to_process = []
    if args.top:
        brands_to_process = all_brands[:args.top]
    elif args.brands:
        for req in args.brands:
            match = next((b for b in all_brands if b.upper() == req.upper()), None)
            if match:
                brands_to_process.append(match)
            else:
                print(f"⚠️  '{req}' not found")
    else:
        parser.print_help()
        return
    
    print(f"\n📊 Generating {len(brands_to_process)} report(s)...\n")
    
    for brand in brands_to_process:
        brand_df = regular_df[regular_df['brand_clean'] == brand]
        if len(brand_df) < 5:
            continue
        
        safe_name = brand.replace('/', '-').replace('\\', '-')[:40]
        output_path = output_folder / f"Brand_Report_{safe_name}.xlsx"
        
        result = create_brand_report(brand_df, brand, output_path, date_range, category_margin_lookup, brand_cat_rev)
        
        icon = "✅" if result['vs_category'] >= 0 else "⚠️"
        rank_str = f"#{result['category_rank']}/{result['category_total']}" if result['category_rank'] > 0 else ""
        print(f"   {icon} {brand}")
        print(f"      ${result['revenue']:,.0f}  |  {result['margin']:.1f}%  |  {result['primary_category']} {rank_str}")
    
    print(f"\n✅ Reports saved to: {output_folder}\n")


if __name__ == "__main__":
    main()
